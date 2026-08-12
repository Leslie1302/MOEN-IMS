"""Turn the three store inventories into inventory bulk-upload templates.

Required columns (from UploadInventoryView): name, quantity, category, code,
unit, warehouse. Category / unit / warehouse must match existing rows in prod
BY EXACT NAME or the row is rejected.

Every row carries a `match` column:
    exact    — description maps unambiguously onto the canonical list
    inferred — I picked between plausible codes; CHECK THESE
    NEW      — not on the canonical list; code minted on your convention
"""
from openpyxl import Workbook, load_workbook

# Category prefix -> the exact Category name that exists in prod.
CAT = {
 'SMA':'Stay & Support Hardware','SMS':'Stay & Support Hardware','SM':'Stay & Support Hardware',
 'IPA':'Insulators','BNW':'Bolts, Nuts & Washers',
 'LHP':'Poles','LHPC':'Poles','LHPC/A':'Poles',
 'CA':'Crossarms','BTC':'Connectors & Clamps','CL':'Connectors & Clamps',
 'DT':'Transformers','SE':'Switchgear & Protection','SCA':'Switchgear & Protection',
 'LFU':'LV Fuse Equipment','EM':'Earthing & Grounding','CCA':'Cables & Conductors',
 'SCM':'Service Connection Materials',
 'EMSP':'Metering Equipment','EMTP':'Metering Equipment','CM':'Metering Equipment',
 'SLA':'Street Lighting','LB':'Lamps & Bulbs','SL':'Solar Equipment',
}
def category_for(code):
    for p in sorted(CAT, key=len, reverse=True):
        if code.upper().startswith(p.upper()): return CAT[p]
    return 'Stay & Support Hardware'

UNIT = {'SINGLE':'pcs','EA':'pcs','METER':'m','METERS':'m','M':'m','SET':'set','PR':'pair',None:'pcs'}

# description -> (canonical name, code, match)
M = {
 # ---- poles
 '10m steel tubular pole':('10M Steel Tubular Pole','LHP007','exact'),
 '10m wooden pole':('10m Wooden Pole','LHP004','exact'),
 '9m wooden pole':('9m Wooden Pole','LHP003','exact'),
 '7m wooden poles':('7m Wood Pole','LHP001','exact'),
 '8m wooden poles':('8m Wood Pole','LHP002','exact'),
 '11m tubular steel pole':('11m Steel Tubular Pole','LHP008','exact'),
 '10m steel tubular pole ':('10M Steel Tubular Pole','LHP007','exact'),
 '10m ht poles':('10 Meter Concrete Pole','LHP010','inferred'),
 '11m steel tubular pole':('11m Steel Tubular Pole','LHP008','exact'),
 '7m service pole':('7m Wood Pole','LHP001','inferred'),
 'wooden stay block':('Wooden stay block','LHP009','exact'),
 'stay wooden block':('Wooden stay block','LHP009','exact'),
 # ---- insulators
 '33kv pin insulators(polymeric)':('33KV Pin Insulator - Polymer','IPA003','exact'),
 '33kv pin insualtor (porceilain)':('33KV Pin Insulator- Porcelain','IPA004','exact'),
 '33kv pin insulator c/w spindle':('33KV Pin Insulator c/w Spindle','IPA005','exact'),
 '11kv pin insulator':('11KV Pin Insulator c/w Spindle - Porcelain','IPA001','inferred'),
 '11kv pin insulators':('11KV Pin Insulator c/w Spindle - Porcelain','IPA001','inferred'),
 '33kv binding stirrup pairs':('Binding Stirrup - 33KV','IPA008','exact'),
 '11kv binding stirrup pairs':('Binding Stirrup - 11KV','IPA007','exact'),
 '33kv binding stirrup':('Binding Stirrup - 33KV','IPA008','exact'),
 '11kv binding stirrup':('Binding Stirrup - 11KV','IPA007','exact'),
 'alu binding wire (mtrs)':('Binding Wire - Aluminium','IPA006','exact'),
 '33kv pin spindle':('33KV Pin Insulator c/w Spindle','IPA005','inferred'),
 'composite suspension (strain insulator polymeric)':('33KV Strain Insulators (Polymer)','IPA014','inferred'),
 # ---- crossarms
 '11kv angle iron crossarm':('11KV Angle Iron Crossarm','CA006','inferred'),
 '33kv angle iron crossarm':('33KV Angle Iron & Arms only','CA004','inferred'),
 '33kv angle iron straps':('33KV Angle Iron Crosssarm c/w tie Straps','CA001','inferred'),
 '33kv channel iron crossarm':('Channel Iron Crossarm','CA005','exact'),
 'channel iron 33kv':('Channel Iron Crossarm','CA005','exact'),
 '33kv angle iron':('33KV Angle Iron & Arms only','CA004','inferred'),
 '11kv angle iron':('11KV Angle Iron Crossarm','CA006','inferred'),
 'crossarms brackets (11kv)':('11KV Angle Iron Crossarm','CA006','inferred'),
 'crossam bracket (33kv)':('33KV Angle Iron & Arms only','CA004','inferred'),
 # ---- bolts
 'm20 x 40':('M20 x 40','BNW007','exact'),
 'm20 x 280':('M20 x280','BNW009','exact'),
 'm16 & 280':('M16 x 280','BNW005','exact'),
 'm16 & 220':('M16 x 220','BNW002','exact'),
 'm16 x 260/300':('M16 x 260','BNW004','inferred'),
 'm16x300':('M16 x 300','BNW006','exact'),
 'm22 x 40 bolts $ nuts':('M22 x 40','BNW010','exact'),
 'm16 x 240 bolts $ nuts':('M16 x 240','BNW003','exact'),
 'm16 x40 bolts and nuts':('M16 x 40','BNW001','exact'),
 'm20 x 40 bolts $ nuts':('M20 x 40','BNW007','exact'),
 'm16 x 220 bolts $ nuts':('M16 x 220','BNW002','exact'),
 'm20 x 280mm bolts, nuts $ wash':('M20 x280','BNW009','exact'),
 'bolts and nuts (m20 x 40)':('M20 x 40','BNW007','exact'),
 'bolts and nuts (m20 x 280)':('M20 x280','BNW009','exact'),
 'm16 curved square washers':('M16 Curved Square Washer','BNW016','NEW'),
 'm20 curved square washers':('M20 Curved Square Washer','BNW017','NEW'),
 # ---- connectors
 'alu clamp for 120sqmm. cond.':('Al Clamps for 120sqmm','BTC020','exact'),
 'al. clamp for 120sqmm pg':('Al Clamps for 120sqmm','BTC020','exact'),
 'al clamps for 120mm2':('Al Clamps for 120sqmm','BTC020','exact'),
 'alu line tap 50/50sqmm.':('Aluminium Line tap','BTC019','exact'),
 'alu line tap':('Aluminium Line tap','BTC019','exact'),
 'copper lug for 70sqmm':('70sqmm Cable Lug-Cu','BTC014','exact'),
 '35mm sq cu cable lugs':('35sqmm Cable Lug-Cu','BTC013','exact'),
 '16cu/120 alu bimetallic':('16Cu/120Al Bimetallic Connector','BTC027','NEW'),
 'bimetalic connector':('16Cu/50AL Bimetallic Connector','BTC010','inferred'),
 'strain clamps':('Tension Clamps','EM003','inferred'),
 'strain clamps for 120mm2':('Tension Clamps','EM003','inferred'),
 'clevis thimble':('Clevis Thimble Sets Complete CAB100 incl. 2.5','SMA007','exact'),
 'celvis thimble set':('Clevis Thimble Sets Complete CAB100 incl. 2.5','SMA007','exact'),
 'clevis socket tongue':('Clevis Socket Tongue','SMA017','NEW'),
 'socket tongue':('Clevis Socket Tongue','SMA017','NEW'),
 'clevis ball hook':('Ball Hooks','SMA008','exact'),
 'ball ended':('Ball Hooks','SMA008','inferred'),
 'tap off clamp connector for 120 x 120sqmm':('Tap Off Clamp Connector for 120 X 120sqmm','BTC026','exact'),
 # ---- transformers
 '33/0.433.50kva 3-ph':('33/50KVA, 3-ph','DT33/3PH006','NEW'),
 '33/0/433, 100kva 3-ph':('33/100KVA, 3-ph','DT33/3PH001','exact'),
 '33/0.433, 200kva 3-ph':('33/200KVA, 3-ph','DT33/3PH002','exact'),
 '11/0.433, 200kva 3-ph':('11/200KVA, 3-ph','DT11/3PH002','exact'),
 '11/0.433, 315kva 3-ph':('11/315KVA, 3-ph','DT11/3PH003','exact'),
 '11/0.433, 50kva-3-ph':('11/50KVA, 3-ph','DT11/3PH004','NEW'),
 '11/0.433, 100kva 3-ph':('11/100KVA, 3-ph','DT11/3PH001','exact'),
 '33/0.415 ,225kva':('33/225KVA, 3-ph','DT33/3PH003','exact'),
 '20/0.25, 50kva 1-ph shield wire':('20/50KVA, 1-ph Shield Wire','DT20/1PH001','NEW'),
 '11.5/0.433, 200kva transformer':('11/200KVA, 3-ph','DT11/3PH002','inferred'),
 '11.5/0.433, 100kva transformer':('11/100KVA, 3-ph','DT11/3PH001','inferred'),
 '33.5/0.433, 100kva transformer':('33/100KVA, 3-ph','DT33/3PH001','inferred'),
 '33.5/0.433, 200kva transformer':('33/200KVA, 3-ph','DT33/3PH002','inferred'),
 '11.5/0.433, 50kva transformer':('11/50KVA, 3-ph','DT11/3PH004','NEW'),
 '33.5/0.433, 50kva transformer':('33/50KVA, 3-ph','DT33/3PH006','NEW'),
 '50kva transformer 11.5/0.433kv':('11/50KVA, 3-ph','DT11/3PH004','NEW'),
 '100kva 33/0.433 transformer':('33/100KVA, 3-ph','DT33/3PH001','inferred'),
 '50kva tranformer (shakthi)':('33/50KVA, 3-ph','DT33/3PH006','inferred'),
 '50kva, 11.5kv transformer (pawell)':('11/50KVA, 3-ph','DT11/3PH004','inferred'),
 '50kva, 33.5kv transformer (recons)':('33/50KVA, 3-ph','DT33/3PH006','inferred'),
 '50kva, 33.5kv transformer':('33/50KVA, 3-ph','DT33/3PH006','inferred'),
 # ---- conductors
 '120sqmm hd alu cond. (mtrs.)':('120sqmm HD ALU Conductor','CCA016','exact'),
 '120sqmm hd alu conductor':('120sqmm HD ALU Conductor','CCA016','exact'),
 '50sqmm hd alu cond. (mtrs.)':('50sqmm HD ALU Conductor','CCA017','exact'),
 '50sqmm hd alu conductor':('50sqmm HD ALU Conductor','CCA017','exact'),
 '70sqmm pvc cu cond. (mtrs)':('70sqmm PVC Cu Conductor','CCA021','NEW'),
 '25sqmm pvc conductor':('25sqmm PVC Insulated Cable-Alu','CCA006','exact'),
 '25sqmm pvc al conductor':('25sqmm PVC Insulated Cable-Alu','CCA006','exact'),
 '25mm2 pvc al. conductor':('25sqmm PVC Insulated Cable-Alu','CCA006','exact'),
 '35sqmm cu pvc insulated':('35sqmm PVC Insulated Cable-Cu','CCA009','exact'),
 '35sqmm pvc cu cond, (mtrs)':('35sqmm PVC Cu Conductor','CCA10','exact'),
 '16sqmm pvc cu conductor':('16sqmm PVC Cu Conductor','CCA004','inferred'),
 '16sqmm pvc cu cond. (mtrs)':('16sqmm PVC Cu Conductor','CCA004','inferred'),
 '35sqmm bare cu cond. (mtrs)':('35sqmm HD Bare Stranded Conductor-Cu','CCA007','exact'),
 'earth guard pcv cond. (mtrs)':('Earth Guard PVC Conductor','CCA022','NEW'),
 '10sqmm pvc cu conductor (mtrs)':('10sqmm PVC Cu Conductor','CCA023','NEW'),
 'u/g (4x50mm2)':('4x50sqmm Underground Cable-Alu','CCA024','NEW'),
 'dead end preformed':('Dead End Preformed','SMA018','NEW'),
 'pvc pipes':('PVC Pipe','CCA025','NEW'),
 # ---- switchgear
 '11kv load isolators':('11KV Load isolator c/w Acc - Polymer','SE002','inferred'),
 '33kv dropout fuse gear':('33KV Dropout Fusegear - Polymer','SE013','inferred'),
 '11kv dropout fuse gear':('11KV Dropout Fusegear - Polymer','SE008','inferred'),
 '11kv drop-out fuse gear':('11KV Dropout Fusegear - Polymer','SE008','inferred'),
 '33kv fuse link 5 amp':('33KV Expulsion Fuselink - 5A','SE018','exact'),
 '33kv fuse 3 amp':('33KV Expulsion Fuselink - 3A','SE017','exact'),
 '11kv fuse link 5 amp':('11KV Expulsion Fuselink - 5A','SE021','exact'),
 '11kv fuse link 5a':('11KV Expulsion Fuselink - 5A','SE021','exact'),
 '11kv fuse link 3 amp':('11KV Expulsion Fuselink - 3A','SE022','exact'),
 '11kv fuse link 2.5a':('11KV Expulsion Fuselink - 2.5A','SE023','exact'),
 '33kv surge divertor':('33KV Surge Arrester - Polymer','SE015','inferred'),
 '33kv surge divertor/arrestor':('33KV Surge Arrester - Polymer','SE015','inferred'),
 'surge divertor 33kv':('33KV Surge Arrester - Polymer','SE015','inferred'),
 '11kv surge divertor':('11Kv Surge Divertor','SE026','exact'),
 '11kv surge divertor/arrestor':('11Kv Surge Divertor','SE026','exact'),
 'surge divertor 11kv':('11Kv Surge Divertor','SE026','exact'),
 'transformer bracket':('Transformer Mounting Bracket','SE027','NEW'),
 'safety switch 600a':('Safety Switch 600A','SE028','NEW'),
 # ---- LV fuse
 'pole mounted l.v fuse unit (mpc400)':('LV Aerial Fuse Unit-Pole Mounted','LFU001','exact'),
 'pole mounted lv fuse unit mpc400':('LV Aerial Fuse Unit-Pole Mounted','LFU001','exact'),
 'pole mounted l.v fuse unit (pc400)':('LV Aerial Fuse Unit-Pole Mounted','LFU001','exact'),
 'lv fuse 300 amp':('LV Fuse - 300 Amps','LFU010','exact'),
 'lv fuse 200 amp':('LV Fuse - 200 Amps','LFU009','exact'),
 'lv fuse 100 amp':('LV Fuse - 100 Amps','LFU008','exact'),
 'lv fuse 200a':('LV Fuse - 200 Amps','LFU009','exact'),
 'lv fuse 100a':('LV Fuse - 100 Amps','LFU008','exact'),
 'lv fuse 63a':('LV Fuse - 63 Amps','LFU012','exact'),
 'lv fuse 315a':('LV Fuse - 315 Amps','LFU014','NEW'),
 # ---- earthing
 'earth rod complete':('Earthrod c/w Clamps','EM001','exact'),
 'earth rods':('Earthrod c/w Clamps','EM001','inferred'),
 'earth rod only':('Earthrod c/w Clamps','EM001','inferred'),
 'earth rod clamp':('Earthrod Clamps','EM002','exact'),
 # ---- service connection
 'shackle insulators':('Shackle Insulators','SCM001','exact'),
 'shackles insulator':('Shackle Insulators','SCM001','exact'),
 'shackle':('Shackle Insulators','SCM001','inferred'),
 'd-iron bracket complete':('D-Bracket Complete','SCM004','exact'),
 'd-iron/d-bracket c/w pin':('D-Bracket Complete','SCM004','exact'),
 'd-iron extension':('D-Iron Bracket Extension','SCM005','exact'),
 'd-iron extension pins':('D-Iron Bracket Extension Pin','SCM014','NEW'),
 'won piece insulators':('Coachscrew (Won Piece) Insulators','SCM003','exact'),
 'won piece insulator':('Coachscrew (Won Piece) Insulators','SCM003','exact'),
 'won pins/coach screws':('Coachscrew (Won Piece) Insulators','SCM003','exact'),
 '60amp service cutout 1-ph comp.':('Service Cut-Out c/w Fuse - 60A','SCM007','exact'),
 '30amp service cutout 1-ph comp.':('Service Cut-Out c/w Fuse - 30A','SCM006','exact'),
 'service cut-out 30a':('Service Cut-Out c/w Fuse - 30A','SCM006','exact'),
 # ---- stay
 'stay bow':('Stay Bow & Crosshead','SMA011','exact'),
 'stay bow and crosshead':('Stay Bow & Crosshead','SMA011','exact'),
 'guy grip':('Guy Grip (Preform) 50sqmm','SMA004','inferred'),
 'stay guy grip (preform)':('Guy Grip (Preform) 50sqmm','SMA004','inferred'),
 'stay guy grip 50mm preform':('Guy Grip (Preform) 50sqmm','SMA004','exact'),
 'stay wire (mtrs)':('Stay Wire','SMA015','exact'),
 'stay wire':('Stay Wire','SMA015','exact'),
 'stay rod':('Stay Rod','SMA013','exact'),
 'stay thimble':('Stay Thimble','SMA016','NEW'),
 'stay thmble':('Stay Thimble','SMA016','NEW'),
 'stay top bracket':('Stay Pole Bracket','SMA012','inferred'),
 'stay pole bracket':('Stay Pole Bracket','SMA012','exact'),
 'stay plate':('Stay Base Plate','SMA010','exact'),
 'stay base plate':('Stay Base Plate','SMA010','exact'),
 'lv stay insulators':('LV Stay Insulator','SMA001','exact'),
 'stay insulator 11kv (lv)':('LV Stay Insulator','SMA001','exact'),
 'stay insulator 11kv':('LV Stay Insulator','SMA001','exact'),
 'stay insulator 33kv (ht)':('HT Stay Insulator','SMA002','exact'),
 'stay insulator 33kv':('HT Stay Insulator','SMA002','exact'),
 'sectional straps':('Sectional Straps','SMA009','exact'),
 'sectional strains (straps)':('Sectional Straps','SMA009','exact'),
 'strut stay equipment complt':('Strut Stay Equipment Complete','SMS002','exact'),
 'stay equipment complete (set)':('Stay Equipment c/w Accessories','SMS001','exact'),
 # ---- meters
 '5/25 amp energy meters 1-ph':('5/25Amps','EMSP001','exact'),
 '20/80 amp energy meters 3-ph':('20/80Amps 3-PH Meters','EMTP001','exact'),
 'energy meter 3-ph 20/80amp':('20/80Amps 3-PH Meters','EMTP001','exact'),
 'energy meters single phase':('1ph 4wire Credit meter','EMSP007','inferred'),
 'energy meters three phase':('3ph 4wire Credit Meter','EMTP003','inferred'),
 '3-phase credit meter':('Three Phase','CM002','exact'),
 '1-phase credit meter':('Single Phase','CM001','exact'),
}

SECTION = {'bolts & nuts','h-pole ht section','transformers','conductors','lv materials','serivce material'}

def norm(s): return ' '.join(str(s).strip().lower().split())
def qty(v):
    if v is None: return 0
    if isinstance(v,(int,float)): return int(v)
    t=str(v).strip().upper()
    return 0 if t in ('NIL','-','') else int(float(t)) if t.replace('.','',1).isdigit() else 0

def build(store, warehouse, records, out):
    hdr=['name','quantity','category','code','unit','warehouse','match','source_description']
    wb=Workbook(); ws=wb.active; ws.title='stock'
    ws.append(hdr)
    seen={}; unmapped=[]
    for desc, unit, q in records:
        k=norm(desc)
        if k in SECTION or not k: continue
        hit=M.get(k)
        if not hit:
            unmapped.append(desc); continue
        name,code,match=hit
        n=qty(q)
        if code in seen:                      # same item listed twice in one store
            ws.cell(seen[code],2).value += n
            continue
        ws.append([name,n,category_for(code),code,UNIT.get(str(unit).strip().upper() if unit else None,'pcs'),
                   warehouse,match,str(desc).strip()])
        seen[code]=ws.max_row
    for i,h in enumerate(hdr,1):
        ws.column_dimensions[ws.cell(1,i).column_letter].width=max(14,len(h)+4)
    wb.save(out)
    print(f"  {out:<44} {ws.max_row-1:>3} items   unmapped={len(unmapped)}")
    for u in unmapped: print(f"        UNMAPPED: {u}")
    return ws.max_row-1

src=load_workbook('/sessions/sharp-dazzling-mccarthy/mnt/uploads/INVENTORY OF SHEP MATERIALS IN WA STORES  AS AT 01-07-2026.xlsx',data_only=True)['Sheet1']
wa=[(r[1],r[2],r[3]) for r in src.iter_rows(min_row=5,values_only=True) if r[1]]
build('Wa','Wa Area Store',wa,'stock_01_WA.xlsx')

src=load_workbook('/sessions/sharp-dazzling-mccarthy/mnt/uploads/SHEP MATERIAL BALANCES, 2025.xlsx',data_only=True)['Sheet1']
bol=[(r[0],r[1],r[2]) for r in src.iter_rows(min_row=5,values_only=True) if r[0]]
build('Bolga','Bolgatanga Area Store',bol,'stock_02_BOLGA.xlsx')

src=load_workbook('/sessions/sharp-dazzling-mccarthy/mnt/uploads/SHEP STOCK BALANCES CURRENT.xlsx',data_only=True)['Sheet1']
from collections import OrderedDict
last=OrderedDict()
for r in src.iter_rows(min_row=3,values_only=True):
    if r[2] is None: continue
    vals=[v for v in r[3:] if isinstance(v,(int,float))]
    if vals: last[str(r[2]).strip()]=vals[-1]
sun=[(d,'EA',v) for d,v in last.items()]
build('Sunyani','Sunyani Area Store',sun,'stock_03_SUNYANI.xlsx')
