"""
SHEP Community Management Views

This module provides CRUD views for managing SHEP communities and packages,
along with AJAX endpoints for cascading dropdowns and the abbreviation legend page.
"""
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.contrib import messages
from django.views.decorators.http import require_http_methods
import uuid
import random
import string
import io

from .models import (
    SHEPCommunity, Community, MaterialOrder, generate_abbreviation,
    InventoryItem, Warehouse, ProjectType, MemberOfParliament, ProjectConsultant,
)
from .forms import SHEPCommunityForm
from .constants import PROJECT_TYPE_SHEP, PROJECT_TYPE_COST_SHARING, PROJECT_TYPE_STREETLIGHTS
from .services.bulk_import import BulkImportResult, normalize_cell, require_columns


class SuperuserRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Mixin to require superuser access."""
    def test_func(self):
        return self.request.user.is_superuser


class SHEPCommunityListView(SuperuserRequiredMixin, ListView):
    """List all SHEP communities with their packages."""
    model = SHEPCommunity
    template_name = 'Inventory/shep_community_list.html'
    context_object_name = 'communities'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by search query
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(region__icontains=search) |
                Q(district__icontains=search) |
                Q(community__icontains=search) |
                Q(package_number__icontains=search)
            )
        
        # Filter by region
        region = self.request.GET.get('region', '')
        if region:
            queryset = queryset.filter(region=region)
        
        return queryset.order_by('region', 'district', 'community')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['regions'] = SHEPCommunity.objects.values_list('region', flat=True).distinct().order_by('region')
        context['search'] = self.request.GET.get('search', '')
        context['selected_region'] = self.request.GET.get('region', '')
        # Drive the bulk-upload template buttons + project dropdown from the
        # live roster of active project types (not a hardcoded three).
        context['project_types'] = ProjectType.objects.filter(active=True).order_by('sort_order', 'name')
        return context


class SHEPCommunityCreateView(SuperuserRequiredMixin, CreateView):
    """Create a new SHEP community."""
    model = SHEPCommunity
    form_class = SHEPCommunityForm
    template_name = 'Inventory/shep_community_form.html'
    success_url = reverse_lazy('shep_community_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Community created successfully.')
        return super().form_valid(form)


class SHEPCommunityUpdateView(SuperuserRequiredMixin, UpdateView):
    """Update an existing SHEP community."""
    model = SHEPCommunity
    form_class = SHEPCommunityForm
    template_name = 'Inventory/shep_community_form.html'
    success_url = reverse_lazy('shep_community_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Community updated successfully.')
        return super().form_valid(form)


class SHEPCommunityDeleteView(SuperuserRequiredMixin, DeleteView):
    """Delete a SHEP community."""
    model = SHEPCommunity
    template_name = 'Inventory/shep_community_confirm_delete.html'
    success_url = reverse_lazy('shep_community_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Community deleted successfully.')
        return super().delete(request, *args, **kwargs)


class AbbreviationLegendView(LoginRequiredMixin, TemplateView):
    """Display abbreviation legend showing full names and their abbreviations."""
    template_name = 'Inventory/abbreviation_legend.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get unique regions with abbreviations
        regions = SHEPCommunity.objects.values('region', 'region_abbr').distinct().order_by('region')
        
        # Get unique districts with abbreviations
        districts = SHEPCommunity.objects.values('district', 'district_abbr', 'region').distinct().order_by('region', 'district')
        
        # Get unique communities with abbreviations
        communities = SHEPCommunity.objects.values('community', 'community_abbr', 'district', 'region').distinct().order_by('region', 'district', 'community')
        
        context['regions'] = regions
        context['districts'] = districts
        context['communities'] = communities
        return context


# AJAX Endpoints for Cascading Dropdowns

def get_districts_by_region(request):
    """AJAX endpoint to get districts for a selected region."""
    region = request.GET.get('region', '')
    if not region:
        return JsonResponse({'districts': []})
    
    districts = SHEPCommunity.objects.filter(
        region=region,
        is_active=True
    ).values('district', 'district_abbr').distinct().order_by('district')
    
    district_list = [
        {'name': d['district'], 'abbreviation': d['district_abbr'] or generate_abbreviation(d['district'])}
        for d in districts
    ]
    
    return JsonResponse({'districts': district_list})


def get_communities_by_district(request):
    """AJAX endpoint to get communities for a selected district."""
    district = request.GET.get('district', '')
    region = request.GET.get('region', '')
    
    if not district:
        return JsonResponse({'communities': []})
    
    queryset = SHEPCommunity.objects.filter(
        district=district,
        is_active=True
    )
    if region:
        queryset = queryset.filter(region=region)
    
    communities = queryset.values('community', 'community_abbr').distinct().order_by('community')
    
    community_list = [
        {'name': c['community'], 'abbreviation': c['community_abbr'] or generate_abbreviation(c['community'])}
        for c in communities
    ]
    
    return JsonResponse({'communities': community_list})


def get_mps_by_constituency(request):
    """AJAX endpoint to find MP(s) for a given constituency / district / region.

    Used by the community form (and any other form that needs to lock the MP
    selection to a constituency) to auto-populate the MP dropdown. Prevents
    users from picking the wrong MP for a constituency.

    Query params (any combination):
      - constituency: exact-match preferred (case-insensitive)
      - district: fallback when constituency unknown
      - region: further narrowing

    Returns:
      {
        'mps': [{ 'id': 12, 'label': 'Hon. ... — Constituency', 'name': '...',
                  'constituency': '...', 'region': '...', 'district': '...' }, ...],
        'auto_select': <id or null>,  # set when exactly one MP matches
        'exact_match': True/False     # True when matched on constituency, not just region
      }
    """
    constituency = (request.GET.get('constituency') or '').strip()
    district = (request.GET.get('district') or '').strip()
    region = (request.GET.get('region') or '').strip()

    qs = MemberOfParliament.objects.filter(active=True)
    exact_match = False

    if constituency:
        # Case-insensitive exact match on constituency name.
        narrowed = qs.filter(constituency__iexact=constituency)
        if narrowed.exists():
            qs = narrowed
            exact_match = True

    if not exact_match:
        # Fall back to district then region scoping so we at least filter
        # the list down to plausible candidates.
        if district:
            narrowed = qs.filter(district__iexact=district)
            if narrowed.exists():
                qs = narrowed
        if region:
            narrowed = qs.filter(region__iexact=region)
            if narrowed.exists():
                qs = narrowed

    qs = qs.order_by('region', 'constituency', 'name')[:50]
    mps = [
        {
            'id': mp.pk,
            'label': f"{mp.title} {mp.name} — {mp.constituency}",
            'name': mp.name,
            'title': mp.title,
            'constituency': mp.constituency,
            'region': mp.region,
            'district': mp.district,
        }
        for mp in qs
    ]
    return JsonResponse({
        'mps': mps,
        'auto_select': mps[0]['id'] if (exact_match and len(mps) == 1) else None,
        'exact_match': exact_match,
    })


def get_packages_by_community(request):
    """AJAX endpoint to get packages for a selected community (SHEP only)."""
    community = request.GET.get('community', '')
    district = request.GET.get('district', '')
    region = request.GET.get('region', '')
    
    if not community:
        return JsonResponse({'packages': []})
    
    queryset = SHEPCommunity.objects.filter(
        community=community,
        is_active=True
    )
    if district:
        queryset = queryset.filter(district=district)
    if region:
        queryset = queryset.filter(region=region)
    
    packages = queryset.values_list('package_number', flat=True).distinct().order_by('package_number')
    
    return JsonResponse({'packages': list(packages)})


# Prefix per project type for auto-generated package numbers.
# Mirrors the COST/SPEC convention from the interactive form.
_PACKAGE_PREFIX = {
    PROJECT_TYPE_COST_SHARING: 'COST',
    PROJECT_TYPE_STREETLIGHTS: 'STREET',
}


def build_auto_package_number(project_code, district, community, requestor=''):
    """Build an auto package number: PREFIX-DISTRICT-COMMUNITY-REQUESTOR-RANDOM.

    Shared by the interactive AJAX endpoint and the bulk importer so both
    produce identical formats. SHEP keeps human-authored package numbers and
    is never auto-generated here.
    """
    prefix = _PACKAGE_PREFIX.get(project_code, 'SPEC')
    district_abbr = generate_abbreviation(district) if district else 'XX'
    community_abbr = generate_abbreviation(community) if community else 'XX'
    requestor_abbr = generate_abbreviation(requestor) if requestor else 'XX'
    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"{prefix}-{district_abbr}-{community_abbr}-{requestor_abbr}-{random_suffix}"


def generate_auto_package_number(request):
    """
    AJAX endpoint to generate an auto package number for Cost-sharing/Special projects.
    Format: [PREFIX]-[DISTRICT_ABBR]-[COMMUNITY_ABBR]-[REQUESTOR_ABBR]-[RANDOM]
    """
    project_type = request.GET.get('project_type', 'COST')
    district = request.GET.get('district', '')
    community = request.GET.get('community', '')
    requestor = request.GET.get('requestor', '')
    
    # Get prefix based on project type
    prefix = 'COST' if project_type == 'COST' else 'SPEC'
    
    # Get abbreviations
    district_abbr = generate_abbreviation(district) if district else 'XX'
    community_abbr = generate_abbreviation(community) if community else 'XX'
    requestor_abbr = generate_abbreviation(requestor) if requestor else 'XX'
    
    # Generate random suffix
    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    
    # Build package number
    package_number = f"{prefix}-{district_abbr}-{community_abbr}-{requestor_abbr}-{random_suffix}"
    
    return JsonResponse({
        'package_number': package_number,
        'district_abbr': district_abbr,
        'community_abbr': community_abbr,
        'requestor_abbr': requestor_abbr
    })


def download_material_template(request):
    """
    Generate and download an Excel template for bulk material requests.
    Includes new columns: project_type, requestor, community.
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, Fill, PatternFill, Alignment
    except ImportError:
        return HttpResponse(
            "Required packages not installed. Please install pandas and openpyxl.",
            status=500
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Material Request Template"
    
    # Define columns. The two trailing signatory columns are batch-level
    # hints — they're only read from the first non-empty row and map to
    # the memo/letter signatory pickers on the release-letter step.
    columns = [
        'name',
        'quantity',
        'project_type',
        'requestor',
        'region',
        'district',
        'community',
        'consultant',
        'contractor',
        'package_number',
        'warehouse',
        'memo_signatory_title',
        'letter_signatory_title',
    ]
    
    # Add header row
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    column_widths = {
        'A': 30,  # name
        'B': 12,  # quantity
        'C': 15,  # project_type
        'D': 25,  # requestor
        'E': 20,  # region
        'F': 20,  # district
        'G': 20,  # community
        'H': 25,  # consultant
        'I': 25,  # contractor
        'J': 25,  # package_number
        'K': 20,  # warehouse
        'L': 32,  # memo_signatory_title
        'M': 32,  # letter_signatory_title
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data validation for project_type
    project_type_dv = DataValidation(
        type="list",
        formula1='"SHEP,COST,SPEC"',
        showDropDown=False,
        allow_blank=False
    )
    project_type_dv.error = "Please select a valid project type: SHEP, COST, or SPEC"
    project_type_dv.errorTitle = "Invalid Project Type"
    project_type_dv.prompt = "Select project type"
    project_type_dv.promptTitle = "Project Type"
    ws.add_data_validation(project_type_dv)
    project_type_dv.add('C2:C1000')
    
    # Get available materials for reference
    materials = list(InventoryItem.objects.values_list('name', flat=True).order_by('name')[:100])
    if materials:
        # Add materials as reference in a separate sheet
        materials_sheet = wb.create_sheet("Materials Reference")
        materials_sheet.cell(row=1, column=1, value="Available Materials").font = Font(bold=True)
        for idx, material in enumerate(materials, 2):
            materials_sheet.cell(row=idx, column=1, value=material)
        materials_sheet.column_dimensions['A'].width = 40
    
    # Get available warehouses for reference
    warehouses = list(Warehouse.objects.values_list('name', flat=True).order_by('name'))
    if warehouses:
        if 'Materials Reference' in wb.sheetnames:
            ref_sheet = wb['Materials Reference']
        else:
            ref_sheet = wb.create_sheet("Reference")
        
        ref_sheet.cell(row=1, column=3, value="Available Warehouses").font = Font(bold=True)
        for idx, warehouse in enumerate(warehouses, 2):
            ref_sheet.cell(row=idx, column=3, value=warehouse)
        ref_sheet.column_dimensions['C'].width = 30
    
    # Add regions/districts/communities reference
    regions = SHEPCommunity.objects.filter(is_active=True).values('region', 'region_abbr').distinct().order_by('region')[:50]
    if regions:
        if 'Materials Reference' in wb.sheetnames:
            ref_sheet = wb['Materials Reference']
        else:
            ref_sheet = wb.create_sheet("Reference")
        
        ref_sheet.cell(row=1, column=5, value="Regions (Abbreviation)").font = Font(bold=True)
        for idx, r in enumerate(regions, 2):
            ref_sheet.cell(row=idx, column=5, value=f"{r['region']} ({r['region_abbr']})")
        ref_sheet.column_dimensions['E'].width = 35
    
    # Add instructions sheet
    instructions_sheet = wb.create_sheet("Instructions")
    instructions = [
        "BULK MATERIAL REQUEST TEMPLATE - INSTRUCTIONS",
        "",
        "REQUIRED FIELDS:",
        "• name: Material name (must match inventory exactly)",
        "• quantity: Numeric quantity requested",
        "• project_type: SHEP, COST, or SPEC",
        "• requestor: Person/factory/institute making the request (required for COST/SPEC)",
        "• region: Project region",
        "• district: Project district",
        "• community: Project community",
        "",
        "OPTIONAL FIELDS:",
        "• consultant: Project consultant",
        "• contractor: Project contractor",
        "• package_number: Required for SHEP, auto-generated for COST/SPEC",
        "• warehouse: Target warehouse",
        "",
        "SIGNATORY COLUMNS (batch-level, first row only):",
        "• memo_signatory_title: Title flagged for the approval memo (e.g. 'Ag. Director, Power')",
        "• letter_signatory_title: Title flagged for the release letter (e.g. 'Chief Director')",
        "• Leave blank to use the active default from the Signatory admin.",
        "• The picker on the upload form overrides whatever's in the Excel.",
        "",
        "PROJECT TYPES:",
        "• SHEP: Regular SHEP project - select package from dropdown",
        "• COST: Cost-sharing project - package auto-generated as COST-[DIST]-[COMM]-[REQ]-[RANDOM]",
        "• SPEC: Special/other project - package auto-generated as SPEC-[DIST]-[COMM]-[REQ]-[RANDOM]",
        "",
        "NOTES:",
        "• Reference the 'Materials Reference' sheet for valid material names",
        "• Abbreviations are shown in parentheses for regions/districts/communities",
    ]
    
    for idx, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=idx, column=1, value=instruction)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif instruction.endswith(':'):
            cell.font = Font(bold=True)
    
    instructions_sheet.column_dimensions['A'].width = 80
    
    # Add example row
    example_data = [
        'Cement Bag 50kg',
        100,
        'SHEP',
        'John Doe',
        'Greater Accra',
        'Accra Metropolitan',
        'Osu',
        'ABC Consulting',
        'XYZ Construction',
        'SHEP-PKG-001',
        'Main Warehouse',
        'Ag. Director, Power',
        'Chief Director',
    ]
    for col_idx, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="material_request_template.xlsx"'
    
    return response


def _community_template_columns(project_code):
    """
    Per-project column schema for community bulk-import templates.

    SHEP template includes package_number + consultant_name (for explicit
    consultant binding; usually left blank so the region-based resolver
    handles it). MP-routed templates include constituency + mp_name.
    """
    if project_code == PROJECT_TYPE_SHEP:
        return ['region', 'district', 'community', 'package_number', 'consultant_name']
    return ['region', 'district', 'community', 'constituency', 'mp_name']


def download_community_template(request):
    """
    Generate a per-project community bulk-import template.

    Project type is selected via the `project` query parameter:
      ?project=shep            -> SHEP template (with package_number)
      ?project=cost_sharing    -> Cost Sharing template
      ?project=streetlights    -> Streetlights template

    Defaults to SHEP for backward compatibility with the legacy
    `download_shep_community_template` URL.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Required package openpyxl not installed.", status=500)

    project_code = (request.GET.get('project') or PROJECT_TYPE_SHEP).strip().lower()

    try:
        project_type = ProjectType.objects.get(code=project_code, active=True)
    except ProjectType.DoesNotExist:
        return HttpResponse(
            f"Unknown or inactive project type: '{project_code}'. "
            f"Active codes: {', '.join(ProjectType.objects.filter(active=True).values_list('code', flat=True))}.",
            status=400,
        )

    columns = _community_template_columns(project_type.code)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{project_type.name} communities"

    # Theme color per project: SHEP green, Cost Sharing teal, Streetlights amber.
    theme_color = {
        PROJECT_TYPE_SHEP: '2E7D32',
        PROJECT_TYPE_COST_SHARING: '0F6E56',
        PROJECT_TYPE_STREETLIGHTS: 'BA7517',
    }.get(project_type.code, '4F81BD')

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type='solid')

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Per-column widths (approximate, readable defaults).
    width_map = {
        'region': 20, 'district': 25, 'community': 25,
        'package_number': 22, 'constituency': 25, 'mp_name': 30,
        'consultant_name': 30,
    }
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width_map.get(col_name, 20)

    # Example row keyed to the project type so users can see the shape.
    example_by_project = {
        PROJECT_TYPE_SHEP: {
            'region': 'Greater Accra', 'district': 'Accra Metropolitan', 'community': 'Osu',
            'package_number': 'SHEP-PKG-001', 'consultant_name': '',
        },
        PROJECT_TYPE_COST_SHARING: {
            'region': 'Upper West', 'district': 'Lawra Municipal', 'community': 'Eremon',
            'constituency': 'Lawra', 'mp_name': '',
        },
        PROJECT_TYPE_STREETLIGHTS: {
            'region': 'Northern', 'district': 'Tamale Metropolitan', 'community': 'Sagnarigu',
            'constituency': 'Tamale Central', 'mp_name': '',
        },
    }
    example = example_by_project.get(project_type.code, {})
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx, value=example.get(col_name, ''))

    # Instructions sheet, project-specific.
    ins = wb.create_sheet(title="Instructions")
    instructions = [
        f"{project_type.name} community bulk upload",
        "",
        "Required columns (all rows must have these):",
        "  - region",
        "  - district",
        "  - community",
    ]
    if project_type.code == PROJECT_TYPE_SHEP:
        instructions.append("  - package_number  (SHEP-only; required)")
    instructions.extend([
        "",
        "Optional columns:",
    ])
    if project_type.code == PROJECT_TYPE_SHEP:
        instructions.append("  - consultant_name (must match a Project Consultant name on record; blank to use region-based lookup)")
    else:
        instructions.append("  - constituency    (used to look up the MP automatically)")
        instructions.append("  - mp_name         (must match a Member of Parliament name on record; blank to use constituency lookup)")
    instructions.extend([
        "",
        "Notes:",
        "  - Project type is set automatically based on which template you downloaded.",
        "  - Region/district/community abbreviations are auto-generated from the names.",
        "  - Duplicate (region+district+community+package_number+project_type) rows will be skipped.",
        "  - If a name lookup fails, the row will be rejected with a clear error.",
        "",
        "After upload, any rows that fail validation are returned to you as a downloadable error CSV.",
    ])

    for idx, line in enumerate(instructions, 1):
        cell = ins.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = Font(bold=True)
    ins.column_dimensions['A'].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="community_template_{project_type.code}.xlsx"'
    )
    return response


def download_shep_community_template(request):
    """
    Generate and download an Excel template for bulk SHEP community import.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse(
            "Required package openpyxl not installed.",
            status=500
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SHEP Communities"
    
    # Define headers
    headers = ['region', 'district', 'community', 'package_number']
    
    # Style headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add example row
    example_data = ['Greater Accra', 'Accra Metropolitan', 'Osu', 'SHEP-PKG-001']
    for col_idx, value in enumerate(example_data, 1):
        ws.cell(row=2, column=col_idx, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet(title="Instructions")
    instructions = [
        "SHEP Community Bulk Upload Instructions",
        "",
        "Required Columns:",
        "- region: The region name (e.g., 'Greater Accra')",
        "- district: The district name (e.g., 'Accra Metropolitan')",
        "- community: The community name (e.g., 'Osu')",  
        "- package_number: The SHEP package number (e.g., 'SHEP-PKG-001')",
        "",
        "Notes:",
        "- Abbreviations will be auto-generated from the names",
        "- Duplicate entries (same region+district+community) will be skipped",
        "- All fields are required",
    ]
    
    for row_idx, text in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    
    instructions_ws.column_dimensions['A'].width = 60
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="shep_community_template.xlsx"'
    
    return response


def upload_communities(request):
    """
    Project-aware community bulk upload.

    Accepts a multipart POST with:
      - file: the Excel file matching one of the per-project templates
      - project: the project type code (shep / cost_sharing / streetlights)

    Validates each row against the project's column schema, looks up the
    ProjectType FK, optionally binds an MP by name, and creates Community
    rows. Failed rows produce a downloadable error CSV; successful rows
    are committed in a single transaction with the rest skipped on error.

    Permission gate: superuser or Management group.
    """
    from django.shortcuts import redirect
    from django.db import transaction

    if not (request.user.is_superuser or request.user.groups.filter(name='Management').exists()):
        messages.error(request, "You don't have permission to bulk-upload communities.")
        return redirect('community_list')

    if request.method != 'POST':
        return redirect('community_list')

    uploaded_file = request.FILES.get('file')
    project_code = (request.POST.get('project') or '').strip().lower()

    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect('community_list')

    if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect('community_list')

    if not project_code:
        messages.error(request, "Pick a project type before uploading.")
        return redirect('community_list')

    try:
        project_type = ProjectType.objects.get(code=project_code, active=True)
    except ProjectType.DoesNotExist:
        messages.error(request, f"Unknown or inactive project type: '{project_code}'.")
        return redirect('community_list')

    try:
        import pandas as pd
    except ImportError:
        messages.error(request, "Required package pandas not installed.")
        return redirect('community_list')

    try:
        df = pd.read_excel(uploaded_file)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect('community_list')

    expected_cols = _community_template_columns(project_type.code)
    required_for_all = ['region', 'district', 'community']
    if project_type.code == PROJECT_TYPE_SHEP:
        required_for_all = required_for_all + ['package_number']

    missing = require_columns(df, required_for_all)
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}.")
        return redirect('community_list')

    # Normalize column names case-insensitively.
    df.columns = [str(c).strip().lower() for c in df.columns]

    result = BulkImportResult(total_rows=len(df))

    # Pre-fetch active MPs and consultants into name -> instance dicts for fast lookup.
    mp_by_name = {
        mp.name.strip().lower(): mp
        for mp in MemberOfParliament.objects.filter(active=True)
    }
    consultant_by_name = {
        c.name.strip().lower(): c
        for c in ProjectConsultant.objects.filter(active=True)
    }

    rows_to_create = []
    for idx, row in df.iterrows():
        excel_row = idx + 2  # header is row 1; first data row is 2

        region = normalize_cell(row.get('region'))
        district = normalize_cell(row.get('district'))
        community_name = normalize_cell(row.get('community'))
        package_number = normalize_cell(row.get('package_number')) if 'package_number' in df.columns else ''
        constituency = normalize_cell(row.get('constituency')) if 'constituency' in df.columns else ''
        mp_name = normalize_cell(row.get('mp_name')) if 'mp_name' in df.columns else ''
        consultant_name = normalize_cell(row.get('consultant_name')) if 'consultant_name' in df.columns else ''

        # Skip entirely empty rows silently.
        if not (region or district or community_name):
            continue

        # Required-field validation.
        if not region:
            result.add_error(excel_row, 'region', 'Required.', region)
        if not district:
            result.add_error(excel_row, 'district', 'Required.', district)
        if not community_name:
            result.add_error(excel_row, 'community', 'Required.', community_name)
        if project_type.code == PROJECT_TYPE_SHEP and not package_number:
            result.add_error(excel_row, 'package_number', 'Required for SHEP.', package_number)

        # Non-SHEP: auto-generate a package number when the column is blank,
        # matching the interactive form's format. Requires region/district/
        # community to be present (validated above), so only generate for
        # otherwise-valid rows.
        if (project_type.code != PROJECT_TYPE_SHEP and not package_number
                and region and district and community_name):
            package_number = build_auto_package_number(
                project_type.code, district, community_name)

        # MP lookup (optional, MP-routed projects only).
        mp_instance = None
        if mp_name:
            mp_instance = mp_by_name.get(mp_name.lower())
            if mp_instance is None:
                result.add_error(
                    excel_row, 'mp_name',
                    f"No active MP named '{mp_name}' on record. Add the MP first or leave blank.",
                    mp_name,
                )

        # Consultant lookup (optional, SHEP only).
        consultant_instance = None
        if consultant_name:
            consultant_instance = consultant_by_name.get(consultant_name.lower())
            if consultant_instance is None:
                result.add_error(
                    excel_row, 'consultant_name',
                    f"No active consultant named '{consultant_name}' on record. Add the consultant first or leave blank.",
                    consultant_name,
                )

        # If this row has any errors, skip it.
        if any(e.row_number == excel_row for e in result.errors):
            continue

        # Skip duplicates silently.
        if Community.objects.filter(
            region=region, district=district, community=community_name,
            package_number=package_number, project_type=project_type,
        ).exists():
            result.skipped_count += 1
            continue

        rows_to_create.append(Community(
            region=region,
            district=district,
            community=community_name,
            package_number=package_number,
            constituency=constituency,
            member_of_parliament=mp_instance,
            project_consultant=consultant_instance,
            project_type=project_type,
        ))

    # Commit successful rows; if any rows failed, we still commit the good ones
    # but return the error CSV. Schedule officers can fix the bad rows offline.
    if rows_to_create:
        try:
            with transaction.atomic():
                # Use save() one-by-one because the model's save() generates
                # abbreviations; bulk_create skips that.
                for inst in rows_to_create:
                    inst.save()
            result.created_count = len(rows_to_create)
        except Exception as exc:  # noqa: BLE001
            result.add_error(0, '*', f"Database error during commit: {exc}")
            messages.error(request, f"Bulk upload failed: {exc}")
            return redirect('community_list')

    # If errors exist, return the error CSV directly so users can download
    # and fix. Successful rows are persisted before this point.
    if result.has_errors:
        response = HttpResponse(
            result.errors_as_csv(),
            content_type='text/csv',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="community_upload_errors_{project_type.code}.csv"'
        )
        # Surface the summary as a session message that'll show on next page load.
        messages.warning(
            request,
            f"Bulk upload completed with errors: {result.summary()} "
            "Error CSV downloaded — fix the rows and re-upload only those.",
        )
        return response

    # All-clean upload: redirect back with a success message.
    if result.created_count or result.skipped_count:
        messages.success(request, f"Bulk upload: {result.summary()}")
    else:
        messages.warning(request, "Bulk upload: no rows processed (file may be empty).")
    return redirect('community_list')


def download_mp_template(request):
    """Excel template for bulk Member of Parliament import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Required package openpyxl not installed.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Members of Parliament"

    columns = ['title', 'name', 'constituency', 'region', 'district', 'email', 'phone']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='534AB7', end_color='534AB7', fill_type='solid')

    for idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    widths = {'title': 10, 'name': 30, 'constituency': 30, 'region': 18, 'district': 22, 'email': 28, 'phone': 18}
    for idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(col_name, 18)

    example = ['Hon.', 'Mary Asante', 'Ga East', 'Greater Accra', 'Ga East', 'mary.asante@parliament.gh', '+233244000000']
    for idx, value in enumerate(example, 1):
        ws.cell(row=2, column=idx, value=value)

    ins = wb.create_sheet(title="Instructions")
    instructions = [
        "Members of Parliament bulk upload",
        "",
        "Required columns:",
        "  - name",
        "  - constituency",
        "  - region",
        "",
        "Optional columns:",
        "  - title       (defaults to 'Hon.' if blank)",
        "  - district    (used by the consignee resolver as a fallback)",
        "  - email",
        "  - phone",
        "",
        "Duplicate rows (matching name + constituency) will be skipped.",
        "Once uploaded, MPs become available for the consignee resolver and",
        "for explicit binding on community records.",
    ]
    for idx, line in enumerate(instructions, 1):
        cell = ins.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = Font(bold=True)
    ins.column_dimensions['A'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="member_of_parliament_template.xlsx"'
    return response


def upload_members_of_parliament(request):
    """Bulk-import MPs from an uploaded Excel file."""
    from django.shortcuts import redirect
    from django.db import transaction

    if not (request.user.is_superuser or request.user.groups.filter(name='Management').exists()):
        messages.error(request, "You don't have permission to bulk-upload MPs.")
        return redirect('community_list')

    if request.method != 'POST':
        return redirect('community_list')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file or not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect('community_list')

    try:
        import pandas as pd
        df = pd.read_excel(uploaded_file)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect('community_list')

    missing = require_columns(df, ['name', 'constituency', 'region'])
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}.")
        return redirect('community_list')

    df.columns = [str(c).strip().lower() for c in df.columns]
    result = BulkImportResult(total_rows=len(df))
    rows_to_create = []

    for idx, row in df.iterrows():
        excel_row = idx + 2
        name = normalize_cell(row.get('name'))
        constituency = normalize_cell(row.get('constituency'))
        region = normalize_cell(row.get('region'))
        title = normalize_cell(row.get('title')) or 'Hon.'
        district = normalize_cell(row.get('district')) if 'district' in df.columns else ''
        email = normalize_cell(row.get('email')) if 'email' in df.columns else ''
        phone = normalize_cell(row.get('phone')) if 'phone' in df.columns else ''

        if not (name or constituency or region):
            continue
        if not name:
            result.add_error(excel_row, 'name', 'Required.', name)
        if not constituency:
            result.add_error(excel_row, 'constituency', 'Required.', constituency)
        if not region:
            result.add_error(excel_row, 'region', 'Required.', region)

        if any(e.row_number == excel_row for e in result.errors):
            continue

        if MemberOfParliament.objects.filter(name__iexact=name, constituency__iexact=constituency).exists():
            result.skipped_count += 1
            continue

        rows_to_create.append(MemberOfParliament(
            title=title, name=name, constituency=constituency,
            region=region, district=district, email=email, phone=phone,
            active=True,
        ))

    if rows_to_create:
        try:
            with transaction.atomic():
                MemberOfParliament.objects.bulk_create(rows_to_create)
            result.created_count = len(rows_to_create)
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Bulk upload failed: {exc}")
            return redirect('community_list')

    if result.has_errors:
        response = HttpResponse(result.errors_as_csv(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="mp_upload_errors.csv"'
        messages.warning(
            request,
            f"MP bulk upload completed with errors: {result.summary()} "
            "Error CSV downloaded — fix the rows and re-upload only those.",
        )
        return response

    if result.created_count or result.skipped_count:
        messages.success(request, f"MP bulk upload: {result.summary()}")
    else:
        messages.warning(request, "MP bulk upload: no rows processed.")
    return redirect('community_list')


def download_consultant_template(request):
    """Excel template for bulk Project Consultant import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Required package openpyxl not installed.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Consultants"

    columns = ['name', 'firm', 'region', 'district', 'contact_email', 'contact_phone']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')

    for idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    widths = {'name': 30, 'firm': 30, 'region': 18, 'district': 22, 'contact_email': 28, 'contact_phone': 18}
    for idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = widths.get(col_name, 18)

    example = ['Apex Engineering Ltd.', 'Apex Engineering Ltd.', 'Greater Accra', '', 'consultant@apex.com', '+233244000000']
    for idx, value in enumerate(example, 1):
        ws.cell(row=2, column=idx, value=value)

    ins = wb.create_sheet(title="Instructions")
    instructions = [
        "Project Consultants bulk upload",
        "",
        "Required columns:",
        "  - name",
        "  - region   (drives SHEP consignee auto-resolution)",
        "",
        "Optional columns:",
        "  - firm           (engineering firm or consultancy name)",
        "  - district       (narrows binding to specific districts within the region)",
        "  - contact_email",
        "  - contact_phone",
        "",
        "Duplicate (name + region) rows will be skipped.",
        "Once uploaded, consultants become available for the SHEP consignee resolver",
        "and for explicit binding on community records.",
    ]
    for idx, line in enumerate(instructions, 1):
        cell = ins.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = Font(bold=True)
    ins.column_dimensions['A'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="project_consultant_template.xlsx"'
    return response


def upload_project_consultants(request):
    """Bulk-import Project Consultants from an uploaded Excel file."""
    from django.shortcuts import redirect
    from django.db import transaction

    if not (request.user.is_superuser or request.user.groups.filter(name='Management').exists()):
        messages.error(request, "You don't have permission to bulk-upload consultants.")
        return redirect('community_list')

    if request.method != 'POST':
        return redirect('community_list')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file or not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect('community_list')

    try:
        import pandas as pd
        df = pd.read_excel(uploaded_file)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect('community_list')

    missing = require_columns(df, ['name', 'region'])
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}.")
        return redirect('community_list')

    df.columns = [str(c).strip().lower() for c in df.columns]
    result = BulkImportResult(total_rows=len(df))
    rows_to_create = []

    for idx, row in df.iterrows():
        excel_row = idx + 2
        name = normalize_cell(row.get('name'))
        region = normalize_cell(row.get('region'))
        firm = normalize_cell(row.get('firm')) if 'firm' in df.columns else ''
        district = normalize_cell(row.get('district')) if 'district' in df.columns else ''
        email = normalize_cell(row.get('contact_email')) if 'contact_email' in df.columns else ''
        phone = normalize_cell(row.get('contact_phone')) if 'contact_phone' in df.columns else ''

        if not (name or region):
            continue
        if not name:
            result.add_error(excel_row, 'name', 'Required.', name)
        if not region:
            result.add_error(excel_row, 'region', 'Required.', region)

        if any(e.row_number == excel_row for e in result.errors):
            continue

        if ProjectConsultant.objects.filter(name__iexact=name, region__iexact=region).exists():
            result.skipped_count += 1
            continue

        rows_to_create.append(ProjectConsultant(
            name=name, firm=firm, region=region, district=district,
            contact_email=email, contact_phone=phone, active=True,
        ))

    if rows_to_create:
        try:
            with transaction.atomic():
                ProjectConsultant.objects.bulk_create(rows_to_create)
            result.created_count = len(rows_to_create)
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Bulk upload failed: {exc}")
            return redirect('community_list')

    if result.has_errors:
        response = HttpResponse(result.errors_as_csv(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="consultant_upload_errors.csv"'
        messages.warning(
            request,
            f"Consultant bulk upload completed with errors: {result.summary()} "
            "Error CSV downloaded — fix the rows and re-upload only those.",
        )
        return response

    if result.created_count or result.skipped_count:
        messages.success(request, f"Consultant bulk upload: {result.summary()}")
    else:
        messages.warning(request, "Consultant bulk upload: no rows processed.")
    return redirect('community_list')


def upload_shep_communities(request):
    """
    Process bulk SHEP community upload from an Excel file.

    Legacy upload endpoint kept for backward compatibility with bookmarks
    pointing at /upload-shep-communities/. Newer flows use upload_communities
    with ?project=shep, which goes through the proper bulk-import service.
    """
    if not request.user.is_superuser:
        messages.error(request, "You don't have permission to upload communities.")
        return redirect('shep_community_list')

    if request.method != 'POST':
        return redirect('shep_community_list')

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        messages.error(request, "No file uploaded.")
        return redirect('shep_community_list')

    if not uploaded_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect('shep_community_list')

    try:
        import pandas as pd
    except ImportError:
        messages.error(request, "Required package pandas is not installed on the server.")
        return redirect('shep_community_list')

    try:
        df = pd.read_excel(uploaded_file)

        required_columns = ['region', 'district', 'community', 'package_number']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            messages.error(
                request,
                f"Missing required columns: {', '.join(missing_columns)}",
            )
            return redirect('shep_community_list')

        created_count = 0
        skipped_count = 0
        error_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                region = str(row['region']).strip()
                district = str(row['district']).strip()
                community_name = str(row['community']).strip()
                package_number = str(row['package_number']).strip()

                # Skip blank or NaN rows.
                if not region or not district or not community_name or region == 'nan':
                    continue

                if SHEPCommunity.objects.filter(
                    region=region,
                    district=district,
                    community=community_name,
                    package_number=package_number,
                ).exists():
                    skipped_count += 1
                    continue

                SHEPCommunity.objects.create(
                    region=region,
                    district=district,
                    community=community_name,
                    package_number=package_number,
                )
                created_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {idx + 2}: {str(e)}")

        msg_parts = []
        if created_count:
            msg_parts.append(f"{created_count} communities created")
        if skipped_count:
            msg_parts.append(f"{skipped_count} duplicates skipped")
        if error_count:
            msg_parts.append(f"{error_count} errors")

        if created_count:
            messages.success(request, ", ".join(msg_parts) + ".")
        elif skipped_count:
            messages.warning(request, ", ".join(msg_parts) + ".")
        else:
            messages.error(request, "No communities were created. " + ", ".join(msg_parts))

        if errors:
            for err in errors[:5]:
                messages.warning(request, err)
            if len(errors) > 5:
                messages.warning(request, f"... and {len(errors) - 5} more errors.")

    except Exception as e:
        messages.error(request, f"Error processing file: {str(e)}")

    return redirect('shep_community_list')



# ── Stock lookup (Phase S) ────────────────────────────────────────────
# Returns current inventory stock for a given InventoryItem. Used by
# material-request and release forms to surface available stock the moment
# the user picks a material.
from django.http import JsonResponse as _JsonResponse


def stock_for_item(request):
    """AJAX: GET /api/stock/?item_id=N[&warehouse_id=W]

    Without warehouse_id: aggregates available stock across every warehouse
    that holds the same material code, so the requestor sees the system-wide
    pool when they choose "any warehouse". Returns a breakdown so they can
    see exactly where the stock lives.

    With warehouse_id: returns the stock for that specific warehouse only.
    """
    item_id = (request.GET.get('item_id') or '').strip()
    warehouse_id = (request.GET.get('warehouse_id') or '').strip()
    if not item_id:
        return _JsonResponse({'error': 'item_id required'}, status=400)
    try:
        from .models import InventoryItem
        anchor = InventoryItem.objects.select_related('unit', 'warehouse').get(pk=item_id)
    except Exception:
        return _JsonResponse({'error': 'not found'}, status=404)

    # All sibling rows holding the same material code across warehouses.
    siblings_qs = InventoryItem.objects.filter(code=anchor.code).select_related('warehouse')

    if warehouse_id:
        row = siblings_qs.filter(warehouse_id=warehouse_id).first()
        available = (row.quantity if row else 0) or 0
        warehouse_name = row.warehouse.name if (row and row.warehouse_id) else ''
        scope = 'warehouse'
        breakdown = []
    else:
        available = sum((r.quantity or 0) for r in siblings_qs)
        warehouse_name = ''
        scope = 'aggregate'
        breakdown = [
            {
                'warehouse': (r.warehouse.name if r.warehouse_id else 'Unassigned'),
                'quantity': r.quantity or 0,
            }
            for r in siblings_qs
        ]

    if available <= 0:
        status = 'out'
    elif available < 10:
        status = 'low'
    else:
        status = 'available'

    return _JsonResponse({
        'available': available,
        'unit': anchor.unit.name if anchor.unit_id else '',
        'warehouse': warehouse_name,
        'scope': scope,
        'breakdown': breakdown,
        'status': status,
        'item_name': anchor.name,
        'item_code': anchor.code,
    })


# ─────────────────────────────────────────────────────────────────────
# Phase S: Stock visibility on release forms
# API endpoint for AJAX lookup of current inventory stock levels
# ─────────────────────────────────────────────────────────────────────
@require_http_methods(['GET'])
def inventory_stock_api(request):
    """
    AJAX endpoint: GET /api/inventory-stock/?item_id=N

    Returns JSON with current stock information:
    {
        "quantity": <decimal>,
        "unit": "<unit_name>",
        "warehouse": "<warehouse_name>",
        "low_stock_threshold": <int>,
        "status": "Available" | "Low" | "Out" | "Critical"
    }
    """
    item_id = request.GET.get('item_id')
    if not item_id:
        return JsonResponse({'error': 'item_id parameter required'}, status=400)

    try:
        item = InventoryItem.objects.get(pk=item_id)
    except InventoryItem.DoesNotExist:
        return JsonResponse({'error': f'Item {item_id} not found'}, status=404)

    # Determine stock status
    LOW_STOCK_THRESHOLD = 10
    CRITICAL_STOCK_THRESHOLD = 5

    if item.quantity <= 0:
        status = 'Out'
    elif item.quantity <= CRITICAL_STOCK_THRESHOLD:
        status = 'Critical'
    elif item.quantity <= LOW_STOCK_THRESHOLD:
        status = 'Low'
    else:
        status = 'Available'

    return JsonResponse({
        'quantity': float(item.quantity),
        'unit': item.unit.name if item.unit else '',
        'warehouse': item.warehouse.name if item.warehouse else '',
        'low_stock_threshold': LOW_STOCK_THRESHOLD,
        'status': status,
        'item_name': item.name,
        'item_code': item.code,
    })


@require_http_methods(['GET'])
def community_detail_api(request):
    """
    AJAX endpoint: GET /api/community-detail/?community_id=N

    Returns comprehensive community details for progressive disclosure:
    {
        "community": "<name>",
        "completion_percent": <0-100>,
        "households_connected": <int>,
        "boq_summary": {
            "total_items": <int>,
            "delivered_count": <int>,
            "pending_count": <int>
        },
        "recent_releases": [
            {"code": "REL-...", "material": "...", "quantity": 100, "status": "..."}
        ],
        "recent_receipts": [
            {"date": "2026-05-18", "material": "...", "quantity": 100, "condition": "Good"}
        ],
        "linked_mp": "<name or null>",
        "linked_consultant": "<name or null>"
    }
    """
    from django.db.models import Count, Q, F
    from .models import BillOfQuantity, MaterialTransport, SiteReceipt, ReleaseLetter

    community_id = request.GET.get('community_id')
    if not community_id:
        return JsonResponse({'error': 'community_id parameter required'}, status=400)

    try:
        community = Community.objects.get(pk=community_id)
    except Community.DoesNotExist:
        return JsonResponse({'error': f'Community {community_id} not found'}, status=404)

    try:
        # Get BoQ items for this community/package
        boq_filter = {
            'region': community.region,
            'district': community.district,
            'community': community.community,
        }
        if community.package_number:
            boq_filter['package_number'] = community.package_number

        boq_items = BillOfQuantity.objects.filter(**boq_filter)

        # Calculate BoQ summary
        total_boq = boq_items.count()
        total_contract_qty = sum(item.contract_quantity for item in boq_items)
        total_received_qty = sum(item.quantity_received for item in boq_items)
        delivered_count = boq_items.filter(quantity_received__gt=0).count()
        pending_count = boq_items.filter(quantity_received=0).count()

        # Completion percentage
        completion_percent = 0
        if total_contract_qty > 0:
            completion_percent = int((total_received_qty / total_contract_qty) * 100)

        # Get recent releases (by package number if SHEP)
        releases = ReleaseLetter.objects.filter(
            Q(request_code__icontains=community.package_number) if community.package_number else Q()
        ).order_by('-id')[:3]

        recent_releases = [
            {
                'code': r.code or r.request_code or 'N/A',
                'material_type': r.material_type,
                'total_quantity': float(r.total_quantity),
                'status': r.workflow_status
            }
            for r in releases
        ]

        # Get recent site receipts (by community location)
        site_receipts = SiteReceipt.objects.filter(
            material_transport__material_order__district=community.district,
            material_transport__material_order__region=community.region
        ).order_by('-received_date')[:3]

        recent_receipts = [
            {
                'date': r.received_date.strftime('%Y-%m-%d'),
                'quantity': float(r.received_quantity),
                'condition': r.condition,
                'received_by': r.received_by.get_full_name() if r.received_by else 'Unknown'
            }
            for r in site_receipts
        ]

        # Get linked MP/consultant
        linked_mp = None
        if community.member_of_parliament:
            linked_mp = community.member_of_parliament.display_name

        linked_consultant = None
        if community.project_consultant:
            linked_consultant = community.project_consultant.name

        return JsonResponse({
            'community': community.community,
            'completion_percent': completion_percent,
            'households_connected': 0,  # Placeholder; no model field yet
            'boq_summary': {
                'total_items': total_boq,
                'delivered_count': delivered_count,
                'pending_count': pending_count,
                'total_contract_qty': float(total_contract_qty),
                'total_received_qty': float(total_received_qty)
            },
            'recent_releases': recent_releases,
            'recent_receipts': recent_receipts,
            'linked_mp': linked_mp,
            'linked_consultant': linked_consultant
        })

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching community detail for {community_id}: {e}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
