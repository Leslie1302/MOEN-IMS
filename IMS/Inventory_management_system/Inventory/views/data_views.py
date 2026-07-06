import json
import logging
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.db.models import Q, Sum
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.views.generic import View, ListView, DetailView
from django.http import JsonResponse, HttpResponse
from django.utils import timezone

from Inventory.models import (
    InventoryItem, Category, Unit, Warehouse, 
    BillOfQuantity, ObsoleteMaterial, MaterialOrder
)
from Inventory.forms import (
    ExcelUploadForm, ObsoleteMaterialForm
)
from .main_views import SuperuserOnlyMixin
from Inventory.utils import is_store_officer, is_management, is_superuser

# Configure logger
logger = logging.getLogger(__name__)

class UploadInventoryView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser  # Only allow admins

    def get(self, request):
        form = ExcelUploadForm()
        return render(request, 'Inventory/upload_inventory.html', {'form': form})

    def post(self, request):
        if not self.request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized access'}, status=403)

        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Read Excel File
                df = pd.read_excel(file, engine='openpyxl')

                # Required columns in Excel
                required_columns = ['name', 'quantity', 'category', 'code', 'unit', 'warehouse']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, "Excel file is missing required columns.")
                    return redirect('dashboard')

                # Load category & unit mappings from DB
                category_mapping = {c.name: c.id for c in Category.objects.all()}
                unit_mapping = {u.name: u.id for u in Unit.objects.all()}
                warehouse_mapping = {w.name: w.id for w in Warehouse.objects.all()}

                for index, row in df.iterrows():
                    # Convert category & unit names to IDs
                    category_id = category_mapping.get(row['category'])
                    unit_id = unit_mapping.get(row['unit'])
                    warehouse_id = warehouse_mapping.get(row['warehouse'])

                    if not category_id or not unit_id or not warehouse_id:
                        messages.error(request, f"Error: Invalid category, unit, or warehouse at row {index + 2}")
                        continue

                    item, created = InventoryItem.objects.get_or_create(
                        code=row['code'],
                        defaults={
                            'name': row['name'],
                            'quantity': row['quantity'],
                            'category_id': category_id,
                            'unit_id': unit_id,
                            'warehouse_id': warehouse_id,
                            'user': request.user
                        }
                    )
                    if not created:
                        item.quantity += row['quantity']
                        item.warehouse_id = warehouse_id
                        item.save()

                messages.success(request, "Inventory updated successfully!")
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")

            return redirect('dashboard')

        return render(request, 'Inventory/upload_inventory.html', {'form': form})

class UploadCategoriesAndUnitsView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser  # Only allow admins

    def get(self, request):
        form = ExcelUploadForm()
        return render(request, 'Inventory/upload_categories_units.html', {'form': form})

    def post(self, request):
        if not self.request.user.is_superuser:
            return JsonResponse({'error': 'Unauthorized access'}, status=403)

        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Read Excel File
                df = pd.read_excel(file, engine='openpyxl')

                # Column Mapping: Convert Excel Columns to Match Model Fields
                column_mapping = {
                    'CATEGORY': 'category',
                    'UNIT': 'unit'
                }
                df.rename(columns=column_mapping, inplace=True)

                # Add Categories to Database
                unique_categories = df['category'].dropna().unique()
                for category_name in unique_categories:
                    Category.objects.get_or_create(name=category_name)

                # Add Units to Database
                unique_units = df['unit'].dropna().unique()
                for unit_name in unique_units:
                    Unit.objects.get_or_create(name=unit_name)

                messages.success(request, "Categories and Units uploaded successfully!")
            except Exception as e:
                messages.error(request, f"Error processing file: {e}")

            return redirect('dashboard')

        return render(request, 'Inventory/upload_categories_units.html', {'form': form})


def list_categories(request):
    categories = list(Category.objects.values('id', 'name'))
    return JsonResponse({'categories': categories})


def list_units(request):
    units = list(Unit.objects.values('id', 'name'))
    return JsonResponse({'units': units})


@login_required
def get_boq_data(request):
    """
    Return BOQ data as JSON for populating cascading dropdowns.
    Returns all unique values for region, district, community, consultant, contractor, and package_number.
    """
    try:
        boq_data = {
            'regions': list(BillOfQuantity.objects.values_list('region', flat=True).distinct().order_by('region')),
            'districts': list(BillOfQuantity.objects.values_list('district', flat=True).distinct().order_by('district')),
            'communities': list(BillOfQuantity.objects.values_list('community', flat=True).distinct().order_by('community')),
            'consultants': list(BillOfQuantity.objects.values_list('consultant', flat=True).distinct().order_by('consultant')),
            'contractors': list(BillOfQuantity.objects.values_list('contractor', flat=True).distinct().order_by('contractor')),
            'package_numbers': list(BillOfQuantity.objects.values_list('package_number', flat=True).distinct().order_by('package_number')),
        }
        
        # Filter out None values
        for key in boq_data:
            boq_data[key] = [item for item in boq_data[key] if item]
        
        return JsonResponse(boq_data)
    except Exception as e:
        logger.error(f"Error fetching BOQ data: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


class DownloadBoQTemplateView(LoginRequiredMixin, View):
    """
    Download the Bill of Quantity upload template.

    Two-sheet workbook:
      • "BoQ Template" — header row + sample rows the user overwrites.
      • "Instructions" — column-by-column guide explaining the contract,
         which columns are required, and how project_type is auto-inferred
         from phase / package_number when the column is left blank.
    """

    def get(self, request):
        columns = [
            'region', 'district', 'community', 'project_type', 'phase',
            'consultant', 'contractor', 'package_number',
            'material_description', 'contract_quantity', 'quantity_received',
        ]
        sample_rows = [
            {
                'region': 'Greater Accra', 'district': 'Ga East',
                'community': 'Abokobi', 'project_type': 'SHEP',
                'phase': 'SHEP-4', 'consultant': 'Acme Engineers Ltd',
                'contractor': 'Beta Power Works',
                'package_number': 'SHEP-4-001',
                'material_description': 'Portland Cement',
                'contract_quantity': 1000, 'quantity_received': 0,
            },
            {
                'region': 'Ashanti', 'district': 'Kumasi Metro',
                'community': 'Asokore Mampong', 'project_type': 'Cost Sharing',
                'phase': 'Phase 2', 'consultant': 'Gamma Consultants',
                'contractor': 'Delta Lines',
                'package_number': 'CS-AHA-01',
                'material_description': 'Steel Reinforcement Bars',
                'contract_quantity': 500, 'quantity_received': 120,
            },
            {
                'region': 'Volta', 'district': 'Hohoe',
                'community': 'Likpe Mate', 'project_type': 'Streetlights',
                'phase': '', 'consultant': 'Epsilon Engineering',
                'contractor': 'Zeta Electricals',
                'package_number': 'SL-VR-007',
                'material_description': 'LED Streetlight Pole 8m',
                'contract_quantity': 60, 'quantity_received': 60,
            },
        ]
        df = pd.DataFrame(sample_rows, columns=columns)

        instructions = pd.DataFrame([
            ('region',               'Required', 'Use one of the 16 Ghana regions, e.g. "Greater Accra", "Ashanti".'),
            ('district',             'Required', 'District name as it appears on the contract.'),
            ('community',            'Required', 'Community name. BoQ→site sync keys off this — keep spelling consistent.'),
            ('project_type',         'Optional', 'One of "SHEP", "Cost Sharing", "Streetlights", "Turnkey", "China Water". If blank, inferred from phase, then package_number prefix; falls back to SHEP.'),
            ('phase',                'Optional', 'Phase label, e.g. "SHEP-4", "Phase 2".'),
            ('consultant',           'Required', 'Consultant name. Auto-resolves the SHEP consignee.'),
            ('contractor',           'Required', 'Contractor name.'),
            ('package_number',       'Required', 'Unique package identifier. Prefix can also drive project_type (SHEP-/CS-/SL-/TK-/CW-).'),
            ('material_description', 'Required', 'Must match an existing InventoryItem name exactly. Unmatched rows are rejected on upload.'),
            ('contract_quantity',    'Required', 'Whole or decimal number. Total contracted quantity for this line.'),
            ('quantity_received',    'Optional', 'Defaults to 0. Site receipts update this automatically afterwards; set a non-zero value only when seeding from existing paper records.'),
        ], columns=['Column', 'Status', 'Notes'])

        notes_block = pd.DataFrame([
            ('Uniqueness',  'Rows are upserted on (item_code, package_number, community).'),
            ('Inference',   'project_type is auto-set from phase keywords (SHEP, Cost, Street, Turnkey, China) or package_number prefix when the column is empty.'),
            ('Sites sync',  'Saving a BoQ row now flips the matching ProjectSite status (Planned → Active → Completed) — Ghana map updates automatically.'),
            ('Backfill',    'After first import, run: python manage.py sync_sites_from_boq'),
        ], columns=['Topic', 'Detail'])

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='BoQ Template')
            instructions.to_excel(writer, index=False, sheet_name='Instructions', startrow=0)
            notes_block.to_excel(writer, index=False, sheet_name='Instructions', startrow=len(instructions) + 3)

            # Column widths so the template opens looking presentable.
            from openpyxl.utils import get_column_letter
            wb = writer.book
            template_ws = wb['BoQ Template']
            widths_main = [16, 16, 18, 14, 12, 24, 24, 18, 28, 12, 12]
            for i, w in enumerate(widths_main, 1):
                template_ws.column_dimensions[get_column_letter(i)].width = w
            inst_ws = wb['Instructions']
            for i, w in enumerate([22, 12, 80], 1):
                inst_ws.column_dimensions[get_column_letter(i)].width = w

        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="boq_upload_template.xlsx"'
        return response


class DownloadSampleTemplateView(LoginRequiredMixin, View):
    """View to download sample inventory template"""
    
    def get(self, request):
        # Sample data for the template
        sample_data = {
            'name': [
                'Portland Cement',
                'Steel Reinforcement Bars',
                'Sand',
                'Gravel',
                'Timber Planks'
            ],
            'quantity': [
                1000,
                500,
                2000,
                1500,
                800
            ],
            'category': [
                'Construction Materials',
                'Steel Products',
                'Aggregates',
                'Aggregates',
                'Timber Products'
            ],
            'code': [
                'CEM-001',
                'STEEL-001',
                'SAND-001',
                'GRAVEL-001',
                'TIMBER-001'
            ],
            'unit': [
                'Bags',
                'Tons',
                'Cubic Meters',
                'Cubic Meters',
                'Pieces'
            ],
            'warehouse': [
                'Main Warehouse',
                'Main Warehouse',
                'Northern Regional Warehouse',
                'Western Regional Warehouse',
                'Eastern Regional Warehouse'
            ]
        }
        
        # Create DataFrame
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inventory Template')
        
        output.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sample_inventory_template.xlsx"'
        
        return response


class MaterialLegendView(LoginRequiredMixin, ListView):
    template_name = 'Inventory/material_legend.html'
    context_object_name = 'materials'
    paginate_by = 25  # Show 25 items per page

    def get_queryset(self):
        # Show all materials to all users for transparency
        queryset = InventoryItem.objects.all().order_by('code')
        return queryset.select_related('category', 'unit', 'warehouse')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Material Code Legend'
        return context


class MaterialHeatmapView(LoginRequiredMixin, View):
    template_name = 'Inventory/material_heatmap.html'

    def get(self, request):
        try:
            # Show all orders and items to all authenticated users
            orders = MaterialOrder.objects.all()
            items = InventoryItem.objects.all()

            period = request.GET.get('period', 'month')
            report_type = request.GET.get('type', 'release')

            if report_type in ['release', 'receipt']:
                orders = orders.filter(request_type=report_type.capitalize())
                df = pd.DataFrame.from_records(orders.values('code', 'name', 'processed_quantity', 'date_requested'))
                if df.empty:
                    df = pd.DataFrame(columns=['code', 'name', 'processed_quantity', 'date_requested'])
                    pivot = pd.DataFrame()
                else:
                    df['date'] = pd.to_datetime(df['date_requested'])
                    if period == 'day':
                        df['period'] = df['date'].dt.strftime('%Y-%m-%d')
                    elif period == 'week':
                        df['period'] = df['date'].dt.strftime('Week %U, %Y')
                    else:  # month
                        df['period'] = df['date'].dt.strftime('%b %Y')
                    
                    # Create pivot table with material names for better readability
                    pivot = pd.pivot_table(
                        df, 
                        values='processed_quantity', 
                        index=['code', 'name'], 
                        columns='period', 
                        aggfunc='sum', 
                        fill_value=0
                    )
                    
                    # Reset index to get code and name as columns
                    pivot = pivot.reset_index()
                    
                    # Create a combined label for code and name
                    pivot['material'] = pivot.apply(
                        lambda x: f"{x['code']} - {x['name']}", 
                        axis=1
                    )
                    
                    # Set the combined label as index
                    pivot = pivot.set_index('material').drop(['code', 'name'], axis=1, errors='ignore')
            else:
                # For stock view, show current quantities
                df = pd.DataFrame.from_records(items.values('code', 'name', 'quantity'))
                if not df.empty:
                    df['material'] = df.apply(
                        lambda x: f"{x['code']} - {x['name']}", 
                        axis=1
                    )
                    # Create a DataFrame with material as index and quantity as value
                    pivot = df[['material', 'quantity']].set_index('material').T
                    # Ensure all materials are included as columns
                    for material in df['material'].tolist():
                        if material not in pivot.columns:
                            pivot[material] = 0
                else:
                    pivot = pd.DataFrame()

            # Initialize pivot if empty (for cases with no data)
            if 'pivot' not in locals() or pivot is None:
                pivot = pd.DataFrame()

            # Generate Plotly figure
            if not pivot.empty:
                import plotly.express as px
                
                # Create the heatmap
                fig = px.imshow(
                    pivot.values,
                    labels=dict(x="Time Period", y="Material", color="Quantity"),
                    x=pivot.columns.tolist(),
                    y=pivot.index.tolist(),
                    aspect="auto",
                    color_continuous_scale='Viridis',
                    text_auto=True
                )
                
                # Customize layout
                title = f"{report_type.capitalize()} Heatmap ({period.capitalize()})"
                fig.update_layout(
                    title={
                        'text': title,
                        'y':0.95,
                        'x':0.5,
                        'xanchor': 'center',
                        'yanchor': 'top',
                        'font': {'size': 20, 'family': 'Arial, sans-serif'}
                    },
                    xaxis_title="Time Period",
                    yaxis_title="Material",
                    height=600 + (len(pivot) * 15),  # Adjust height based on number of materials
                    margin=dict(l=150, r=50, b=150, t=100, pad=4),
                    coloraxis_colorbar=dict(
                        title="Quantity",
                        thicknessmode="pixels", thickness=20,
                        lenmode="pixels", len=300,
                        yanchor="top", y=1,
                        ticks="outside"
                    )
                )
                
                # Customize hover text
                fig.update_traces(
                    hovertemplate='<b>Material</b>: %{y}<br>' +
                                 '<b>Period</b>: %{x}<br>' +
                                 '<b>Quantity</b>: %{z}<extra></extra>',
                    texttemplate="%{z}",
                    textfont={"size": 10}
                )
                
                # Convert the plot to HTML
                plot_div = fig.to_html(
                    full_html=False, 
                    include_plotlyjs='cdn',  # Use CDN for Plotly.js
                    config={
                        'displayModeBar': True,
                        'scrollZoom': True,
                        'responsive': True
                    }
                )
                
                has_data = True
            else:
                plot_div = ""
                has_data = False
            
            context = {
                'plot_div': plot_div,
                'report_type': report_type,
                'period': period,
                'has_data': has_data
            }
            return render(request, self.template_name, context)
            
        except Exception as e:
            logger.error(f"Error generating heatmap: {str(e)}", exc_info=True)
            messages.error(request, f"Error generating heatmap: {str(e)}")
            return render(request, self.template_name, {'has_data': False})

    def post(self, request):
        # Show all orders and items to all users for transparency
        orders = MaterialOrder.objects.all()
        items = InventoryItem.objects.all()

        period = request.POST.get('period', 'month')
        report_type = request.POST.get('type', 'release')

        if report_type in ['release', 'receipt']:
            orders = orders.filter(request_type=report_type.capitalize())
            df = pd.DataFrame.from_records(orders.values('code', 'processed_quantity', 'date_requested'))
            if df.empty:
                df = pd.DataFrame(columns=['code', 'processed_quantity', 'date_requested'])
                pivot = pd.DataFrame() # Fixed initialization
            else:
                df['date'] = pd.to_datetime(df['date_requested'])
                if period == 'day':
                    df['period'] = df['date'].dt.date
                elif period == 'week':
                    df['period'] = df['date'].dt.to_period('W').apply(lambda r: r.start_time.date())
                else:  # month
                    df['period'] = df['date'].dt.to_period('M').apply(lambda r: r.start_time.date())
                pivot = pd.pivot_table(df, values='processed_quantity', index='code', columns='period', aggfunc='sum', fill_value=0)
        else:
            df = pd.DataFrame.from_records(items.values('code', 'quantity'))
            if not df.empty:
                pivot = df.set_index('code').T
            else:
                pivot = pd.DataFrame()

        if pivot.empty:
             messages.warning(request, "No data available to generate PDF.")
             return redirect('material_heatmap')

        # Set up a professional, modern style heatmap for PDF
        plt.figure(figsize=(15, 10))
        sns.set_style("whitegrid")
        heatmap = sns.heatmap(
            pivot,
            cmap="Blues",
            annot=True,
            fmt='.0f',
            cbar_kws={'label': 'Quantity', 'shrink': 0.8, 'pad': 0.02, 'orientation': 'vertical'},
            linewidths=0.5,
            annot_kws={'size': 10, 'weight': 'bold'},
            square=True,
            vmin=0,
            vmax=pivot.max().max() * 1.1
        )
        plt.title(f"{report_type.capitalize()} Heatmap ({period.capitalize()})", fontsize=16, pad=20, fontweight='bold', color='#2c3e50')
        plt.xlabel("Time Period", fontsize=12, color='#2c3e50')
        plt.ylabel("Material Code", fontsize=12, color='#2c3e50')

        plt.xticks(rotation=45, ha='right', fontsize=10)
        plt.yticks(fontsize=10)
        plt.tight_layout()

        # Save to PDF buffer
        buffer = BytesIO()
        plt.savefig(buffer, format='pdf', bbox_inches='tight', dpi=300)
        plt.close()
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_heatmap_{period}.pdf"'
        return response

class LowInventorySummaryView(LoginRequiredMixin, ListView):
    """
    View for displaying low inventory items.
    Accessible to all authenticated users.
    Shows all low inventory items regardless of group.
    """
    template_name = 'Inventory/low_inventory_summary.html'
    context_object_name = 'low_items'

    def get_queryset(self):
        """
        Return all low inventory items.
        Accessible to all authenticated users.
        """
        LOW_QUANTITY = 5  # Match the threshold used in Dashboard view
        return InventoryItem.objects.filter(quantity__lte=LOW_QUANTITY).select_related('category', 'unit', 'warehouse').order_by('name')

class BillOfQuantityView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    View for displaying Bill of Quantities.
    Accessible to schedule officers and superusers.
    Shows all BOQ items.
    """
    template_name = 'Inventory/bill_of_quantity.html'
    context_object_name = 'boq_items'
    paginate_by = 50
    paginate_orphans = 5
    
    def test_func(self):
        from Inventory.utils import is_schedule_officer
        user = self.request.user
        return is_schedule_officer(user) or is_store_officer(user) or is_management(user) or is_superuser(user)

    def get_queryset(self):
        """
        Return all BOQ items.
        Accessible to schedule officers and superusers.
        """
        return BillOfQuantity.objects.all().order_by('package_number', 'material_description')

def _resolve_boq_project_type(row):
    """Pick a project_type for a BoQ upload row.

    Priority:
      1. Explicit `project_type` cell in the spreadsheet, normalised
         against the active ProjectType registry (case-insensitive match
         on name OR code; falls back to the verbatim string).
      2. Phase prefix — "SHEP-4", "Shep Phase 2" -> SHEP, etc.
      3. Package-number prefix — "COST-…" -> Cost Sharing,
         "POLES-…" / legacy "STREET-…" -> Poles, "SHEP-…" -> SHEP.
      4. Default 'SHEP' (keeps legacy behaviour when no hint is present).
    """
    from Inventory.models import ProjectType

    registry = {}  # lower-cased name/code -> canonical name
    for t in ProjectType.objects.filter(active=True):
        registry[t.name.lower()] = t.name
        registry[t.code.lower()] = t.name

    # 1. Explicit column.
    raw = row.get('project_type') if hasattr(row, 'get') else None
    if raw is not None and pd.notna(raw):
        s = str(raw).strip()
        if s:
            return registry.get(s.lower(), s)

    # 2. Phase prefix.
    phase = row.get('phase') if hasattr(row, 'get') else None
    if phase is not None and pd.notna(phase):
        p = str(phase).strip().lower()
        if 'shep' in p: return registry.get('shep', 'SHEP')
        if 'cost' in p: return registry.get('cost_sharing', 'Cost Sharing')
        if 'street' in p or 'pole' in p: return registry.get('streetlights', 'Streetlights')
        if 'turnkey' in p: return 'Turnkey'
        if 'china' in p: return 'China Water'

    # 3. Package-number prefix.
    pkg = row.get('package_number') if hasattr(row, 'get') else None
    if pkg is not None and pd.notna(pkg):
        pn = str(pkg).strip().upper()
        if pn.startswith('SHEP'): return registry.get('shep', 'SHEP')
        if pn.startswith('CS') or pn.startswith('COST'): return registry.get('cost_sharing', 'Cost Sharing')
        if pn.startswith('SL') or pn.startswith('STR'): return registry.get('streetlights', 'Streetlights')
        if pn.startswith('TK') or pn.startswith('TURN'): return 'Turnkey'
        if pn.startswith('CW') or pn.startswith('CHINA'): return 'China Water'

    return 'SHEP'


class UploadBillOfQuantityView(LoginRequiredMixin, SuperuserOnlyMixin, View):

    def get(self, request):
        form = ExcelUploadForm()
        return render(request, 'Inventory/upload_bill_of_quantity.html', {'form': form})

    def post(self, request):
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            from Inventory.signals import _boq_bulk
            from django.core.management import call_command
            did_bulk = False
            try:
                df = pd.read_excel(file, engine='openpyxl')
                # material_description must exactly match existing InventoryItem names
                required_columns = [
                    'region', 'district', 'community', 'consultant', 'contractor', 
                    'package_number', 'material_description', 
                    'contract_quantity', 'quantity_received'
                ]
                if not all(col in df.columns for col in required_columns):
                    missing_cols = [col for col in required_columns if col not in df.columns]
                    messages.error(request, f"Excel file is missing required columns: {', '.join(missing_cols)}")
                    return redirect('bill_of_quantity')

                success_count = 0
                error_count = 0
                
                def _cell(value):
                    """Coerce a pandas cell to a clean string. Blank/NaN -> ''.

                    Without this, empty Excel cells arrive as float('nan') and
                    get persisted to CharFields as the literal string 'nan'.
                    """
                    if not pd.notna(value):
                        return ''
                    return str(value).strip()

                # Bulk mode: suppress per-row BoQ signals for the whole loop,
                # then sync sites once in `finally`. Prevents the O(N^2) hang.
                _boq_bulk.on = True
                did_bulk = True
                for index, row in df.iterrows():
                    try:
                        # Handle NaN or empty values
                        contract_qty = int(float(row['contract_quantity'])) if pd.notna(row['contract_quantity']) else 0
                        qty_received = int(float(row['quantity_received'])) if pd.notna(row['quantity_received']) else 0
                        material_description = _cell(row['material_description']) or None
                        
                        if not material_description:
                            logger.warning(f"Row {index + 2}: Missing material_description, skipping")
                            messages.warning(request, f"Row {index + 2}: Missing material description, skipped")
                            error_count += 1
                            continue
                        
                        # Require exact match with existing InventoryItem - no auto-generation allowed
                        inventory_item = InventoryItem.objects.filter(name__iexact=material_description).first()
                        
                        if not inventory_item:
                            # No match found - reject this row to maintain uniformity
                            error_count += 1
                            logger.error(f"Row {index + 2}: Material '{material_description}' not found in inventory system")
                            messages.error(
                                request, 
                                f"Row {index + 2}: Material '{material_description}' does not match any existing inventory item. "
                                f"Please ensure material descriptions exactly match items in your inventory."
                            )
                            continue
                        
                        # Use the matched inventory item's code
                        item_code = inventory_item.code
                        logger.info(f"Row {index + 2}: Matched inventory item '{inventory_item.name}' with code '{item_code}'")

                        # Resolve project type for this row (explicit
                        # column → phase → package prefix → SHEP default).
                        resolved_project_type = _resolve_boq_project_type(row)
                        phase_value = row.get('phase') if 'phase' in row else None
                        if phase_value is not None and not pd.notna(phase_value):
                            phase_value = None

                        boq, created = BillOfQuantity.objects.get_or_create(
                            item_code=item_code,
                            package_number=_cell(row['package_number']),
                            community=_cell(row.get('community')) or None,
                            defaults={
                                'region': _cell(row['region']),
                                'district': _cell(row['district']),
                                'consultant': _cell(row['consultant']),
                                'contractor': _cell(row['contractor']),
                                'material_description': material_description,
                                'contract_quantity': contract_qty,
                                'quantity_received': qty_received,
                                'project_type': resolved_project_type,
                                'phase': phase_value,
                                'user': request.user,
                            }
                        )
                        if not created:
                            # Update existing record
                            boq.region = _cell(row['region'])
                            boq.district = _cell(row['district'])
                            boq.community = _cell(row.get('community')) or None
                            boq.consultant = _cell(row['consultant'])
                            boq.contractor = _cell(row['contractor'])
                            boq.material_description = material_description
                            boq.contract_quantity = contract_qty
                            boq.quantity_received = qty_received
                            boq.project_type = resolved_project_type
                            if phase_value is not None:
                                boq.phase = phase_value
                            boq.save()
                            logger.info(f"Row {index + 2}: Updated existing BOQ item with code '{item_code}' (project_type={resolved_project_type})")
                        else:
                            logger.info(f"Row {index + 2}: Created new BOQ item with code '{item_code}'")
                        
                        success_count += 1

                    except Exception as e:
                        error_count += 1
                        logger.error(f"Error processing row {index + 2}: {str(e)}", exc_info=True)
                        messages.error(request, f"Error at row {index + 2}: {str(e)}")
                        continue  # Skip problematic rows but continue processing others

                if success_count > 0:
                    messages.success(request, f"Bill of Quantity updated successfully! {success_count} items processed.")
                if error_count > 0:
                    messages.warning(request, f"{error_count} rows had errors and were skipped.")
                    
            except Exception as e:
                logger.error(f"Error processing Excel file: {str(e)}", exc_info=True)
                messages.error(request, f"Error processing file: {str(e)}")
            finally:
                # Always clear the bulk flag (worker threads are reused) and,
                # if we imported, recompute site statuses once instead of per row.
                _boq_bulk.on = False
                if did_bulk:
                    try:
                        call_command('sync_sites_from_boq')
                    except Exception as sync_err:
                        logger.error(f"Post-import site sync failed: {sync_err}", exc_info=True)
                    # One summary email to Management instead of one per line item.
                    try:
                        from Inventory.signals import create_notification
                        regions = ', '.join(sorted(str(r) for r in df['region'].dropna().unique()))
                        create_notification(
                            notification_type='boq_updated',
                            title=f'BoQ import: {success_count} lines ({regions})',
                            message=(f'{success_count} Bill of Quantity lines were imported '
                                     f'for {regions}. ({error_count} rows skipped.)'),
                            recipient_group='Management',
                            sender=request.user,
                        )
                    except Exception as note_err:
                        logger.error(f"BoQ summary notification failed: {note_err}", exc_info=True)
            return redirect('bill_of_quantity')

        return render(request, 'Inventory/upload_bill_of_quantity.html', {'form': form})

class ObsoleteMaterialRegisterView(LoginRequiredMixin, View):
    """
    View for registering obsolete materials.
    Displays form with auto-population of material details.
    """
    template_name = 'Inventory/obsolete_material_register.html'
    
    def get(self, request):
        form = ObsoleteMaterialForm()
        
        # Get all inventory items for auto-population in JavaScript
        items = InventoryItem.objects.select_related('category', 'unit', 'warehouse').all()
        inventory_items = list(items.values(
            'id', 'name', 'code', 
            'category__name', 'unit__name', 
            'warehouse__id', 'warehouse__name'
        ))
        
        context = {
            'form': form,
            'inventory_items': json.dumps(inventory_items),
            'title': 'Register Obsolete Material'
        }
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        form = ObsoleteMaterialForm(request.POST)
        
        if form.is_valid():
            obsolete_material = form.save(commit=False)
            obsolete_material.registered_by = request.user
            obsolete_material.save()
            
            messages.success(
                request, 
                f'Successfully registered obsolete material: {obsolete_material.material_name} '
                f'({obsolete_material.quantity} {obsolete_material.unit})'
            )
            
            return redirect('obsolete_material_list')
        
        # If form is invalid, re-render with errors
        items = InventoryItem.objects.select_related('category', 'unit', 'warehouse').all()
        inventory_items = list(items.values(
            'id', 'name', 'code', 
            'category__name', 'unit__name', 
            'warehouse__id', 'warehouse__name'
        ))
        
        context = {
            'form': form,
            'inventory_items': json.dumps(inventory_items),
            'title': 'Register Obsolete Material'
        }
        
        return render(request, self.template_name, context)


class ObsoleteMaterialListView(LoginRequiredMixin, ListView):
    """
    List view for all obsolete materials.
    Displays registered obsolete materials with filtering and search.
    """
    model = ObsoleteMaterial
    template_name = 'Inventory/obsolete_material_list.html'
    context_object_name = 'obsolete_materials'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = ObsoleteMaterial.objects.select_related(
            'material', 'warehouse', 'registered_by', 'reviewed_by'
        ).all()
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by warehouse
        warehouse_id = self.request.GET.get('warehouse')
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        
        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category__icontains=category)
        
        # Search by material name or code
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(material_name__icontains=search) | 
                Q(material_code__icontains=search)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Obsolete Materials Register'
        context['warehouses'] = Warehouse.objects.filter(is_active=True)
        context['status_choices'] = ObsoleteMaterial.STATUS_CHOICES
        
        # Summary statistics
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['total_value'] = queryset.aggregate(
            total=Sum('estimated_value')
        )['total'] or 0
        
        # Count by status
        context['status_counts'] = {}
        for status_code, status_label in ObsoleteMaterial.STATUS_CHOICES:
            context['status_counts'][status_code] = queryset.filter(status=status_code).count()
        
        return context


class ObsoleteMaterialDetailView(LoginRequiredMixin, DetailView):
    """
    Detail view for a single obsolete material record.
    Shows all information including audit trail.
    """
    model = ObsoleteMaterial
    template_name = 'Inventory/obsolete_material_detail.html'
    context_object_name = 'obsolete_material'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Obsolete Material: {self.object.material_name}'
        context['can_review'] = self.request.user.has_perm('Inventory.can_review_obsolete_material')
        context['can_approve'] = self.request.user.has_perm('Inventory.can_approve_disposal')
        return context


@login_required
@permission_required('Inventory.can_review_obsolete_material', raise_exception=True)
def update_obsolete_material_status(request, pk):
    """
    Update the status of an obsolete material record.
    Requires review permission.
    """
    obsolete_material = get_object_or_404(ObsoleteMaterial, pk=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        review_notes = request.POST.get('review_notes', '')
        
        if new_status in dict(ObsoleteMaterial.STATUS_CHOICES):
            obsolete_material.status = new_status
            obsolete_material.reviewed_by = request.user
            obsolete_material.review_date = timezone.now()
            
            if review_notes:
                obsolete_material.review_notes = review_notes
            
            # Handle disposal-specific fields
            if new_status == 'Disposed':
                disposal_method = request.POST.get('disposal_method')
                disposal_date = request.POST.get('disposal_date')
                
                if disposal_method:
                    obsolete_material.disposal_method = disposal_method
                if disposal_date:
                    obsolete_material.disposal_date = disposal_date
            
            obsolete_material.save()
            
            messages.success(
                request,
                f'Status updated to "{new_status}" for {obsolete_material.material_name}'
            )
        else:
            messages.error(request, 'Invalid status value')
    
    return redirect('obsolete_material_detail', pk=pk)
