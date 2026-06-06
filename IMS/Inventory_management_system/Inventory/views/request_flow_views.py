"""
Phase C — two-step Material Request flow views.

Step 1: SelectProjectView (GET /request-material/)
        Renders a tiny form with just project_type. On POST, redirects to
        Step 2 with the chosen code in the URL. Also surfaces the bulk
        upload section (per-project templates + upload).

Step 2: RequestMaterialForProjectView (GET /request-material/start/<code>/)
        Renders the project-specific form for the chosen project_type.
        Common fields (material, quantity, community, warehouse, notes)
        plus per-project additions. Consignee is auto-resolved at save
        time from (project_type, community) and rendered as a preview.

Bulk: download_request_template + upload_requests
        Per-project Excel template + project-aware bulk upload that
        creates MaterialOrder rows with consignee auto-resolution.
"""

import io
import logging

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import Http404, HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.generic import View

from ..models import ProjectType, Community, MaterialOrder, InventoryItem, Warehouse, Signatory
from ..forms.request_flow import ProjectSelectorForm, form_class_for_project
from ..constants import (
    active_project_types, project_type_to_charfield,
    PROJECT_TYPE_SHEP, PROJECT_TYPE_COST_SHARING, PROJECT_TYPE_STREETLIGHTS,
)
from ..services.consignee_resolver import resolve_consignee
from ..services.bulk_import import BulkImportResult, normalize_cell, require_columns
from .order_views import generate_request_code  # reuse the existing code generator

logger = logging.getLogger(__name__)


class SelectProjectView(LoginRequiredMixin, View):
    """Step 1 — project picker."""

    template_name = 'Inventory/request_select_project.html'

    def get(self, request):
        form = ProjectSelectorForm()
        return render(request, self.template_name, {
            'form': form,
            'project_types': active_project_types(),
        })

    def post(self, request):
        form = ProjectSelectorForm(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {
                'form': form,
                'project_types': active_project_types(),
            })
        chosen = form.cleaned_data['project_type']
        return redirect('request_material_for_project', project_code=chosen.code)


class RequestMaterialForProjectView(LoginRequiredMixin, View):
    """Step 2 — project-specific request form."""

    template_name = 'Inventory/request_material_v2.html'

    def _get_project_type(self, project_code):
        try:
            return ProjectType.objects.get(code=project_code, active=True)
        except ProjectType.DoesNotExist:
            raise Http404(f"Unknown or inactive project type: {project_code}")

    def _signatory_context(self):
        """Designation-led pickers for the letterhead-override section."""
        active = Signatory.objects.filter(active=True)
        return {
            'memo_signatories': active.filter(
                is_default_for_release_memo=True,
            ).order_by('title'),
            'letter_signatories': active.filter(
                is_default_for_release_letter=True,
            ).order_by('title'),
        }

    def get(self, request, project_code):
        project_type = self._get_project_type(project_code)
        FormClass = form_class_for_project(project_type)
        form = FormClass(project_type_instance=project_type)
        return render(request, self.template_name, {
            'form': form,
            'project_type': project_type,
            **self._signatory_context(),
        })

    def post(self, request, project_code):
        project_type = self._get_project_type(project_code)
        FormClass = form_class_for_project(project_type)
        form = FormClass(request.POST, request.FILES, project_type_instance=project_type)

        if not form.is_valid():
            return render(request, self.template_name, {
                'form': form,
                'project_type': project_type,
                **self._signatory_context(),
            })

        try:
            with transaction.atomic():
                instance = form.save(commit=False, user=request.user)
                # Stamp request_code if the model expects one.
                if hasattr(instance, 'request_code') and not instance.request_code:
                    instance.request_code = generate_request_code()
                if hasattr(instance, 'date_requested') and not instance.date_requested:
                    instance.date_requested = timezone.now()
                if not getattr(instance, 'user_id', None):
                    instance.user = request.user
                instance.save()
        except Exception as exc:  # noqa: BLE001
            logger.exception("Failed to save material order via two-step flow")
            messages.error(request, f"Could not save the request: {exc}")
            return render(request, self.template_name, {
                'form': form,
                'project_type': project_type,
                **self._signatory_context(),
            })

        messages.success(
            request,
            f"Request submitted ({instance.request_code if hasattr(instance, 'request_code') else instance.pk}). "
            f"Project: {project_type.name}. Consignee resolved and recorded.",
        )
        return redirect('material_orders')


def _request_template_columns(project_code):
    """Per-project column schema for material request bulk templates.

    The two trailing signatory columns are batch-level hints — they only
    need to be filled on the first row. They map to the designation
    pickers on the Release Letter generation page so the right officer's
    name appears on the memo and the release letter.
    """
    # package_number is offered for EVERY project type — packages are used
    # across programmes for tracking/BoQ reconciliation. Left blank, it is
    # looked up from the community.
    base = ['material', 'quantity', 'region', 'district', 'community',
            'warehouse', 'notes', 'package_number']
    if project_code == PROJECT_TYPE_COST_SHARING:
        body = base + ['beneficiary_contribution']
    elif project_code == PROJECT_TYPE_STREETLIGHTS:
        body = base + ['pole_height_m', 'lumen_rating', 'pole_type']
    else:
        body = base
    return body + ['memo_signatory_title', 'letter_signatory_title']


def download_request_template(request):
    """
    Per-project material-request bulk template.
    Query: ?project=shep|cost_sharing|streetlights
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Required package openpyxl not installed.", status=500)

    project_code = (request.GET.get('project') or PROJECT_TYPE_SHEP).strip().lower()
    try:
        project_type = ProjectType.objects.get(code=project_code, active=True)
    except ProjectType.DoesNotExist:
        return HttpResponse(f"Unknown or inactive project type: '{project_code}'.", status=400)

    columns = _request_template_columns(project_type.code)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{project_type.name} requests"

    theme_color = {
        PROJECT_TYPE_SHEP: '2E7D32',
        PROJECT_TYPE_COST_SHARING: '0F6E56',
        PROJECT_TYPE_STREETLIGHTS: 'BA7517',
    }.get(project_type.code, '4F81BD')

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type='solid')

    for idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    width_map = {
        'material': 30, 'quantity': 12, 'region': 18, 'district': 22, 'community': 22,
        'warehouse': 20, 'notes': 30, 'package_number': 22,
        'beneficiary_contribution': 35, 'pole_height_m': 14, 'lumen_rating': 14, 'pole_type': 22,
        'memo_signatory_title': 32, 'letter_signatory_title': 32,
    }
    for idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width_map.get(col_name, 18)

    # Example row keyed to the project type. Signatory columns share the
    # same example titles across projects so users learn the convention.
    signatory_example = {
        'memo_signatory_title': 'Ag. Director, Power',
        'letter_signatory_title': 'Chief Director',
    }
    example_by_project = {
        PROJECT_TYPE_SHEP: {
            'material': 'Pole - 11kV concrete', 'quantity': 24,
            'region': 'Greater Accra', 'district': 'Ga East', 'community': 'Abokobi',
            'warehouse': '', 'notes': '', 'package_number': 'SHEP-PKG-024',
            **signatory_example,
        },
        PROJECT_TYPE_COST_SHARING: {
            'material': 'Conductor - ACSR Dog', 'quantity': 5000,
            'region': 'Upper West', 'district': 'Lawra Municipal', 'community': 'Eremon',
            'warehouse': '', 'notes': '',
            'beneficiary_contribution': '30% community contribution agreed at 10 March 2026 meeting',
            **signatory_example,
        },
        PROJECT_TYPE_STREETLIGHTS: {
            'material': 'Streetlight assembly', 'quantity': 50,
            'region': 'Northern', 'district': 'Tamale Metropolitan', 'community': 'Sagnarigu',
            'warehouse': '', 'notes': '',
            'pole_height_m': 8, 'lumen_rating': 12000, 'pole_type': 'galvanised steel, octagonal',
            **signatory_example,
        },
    }
    example = example_by_project.get(project_type.code, dict(signatory_example))
    for idx, col_name in enumerate(columns, 1):
        ws.cell(row=2, column=idx, value=example.get(col_name, ''))

    # Instructions sheet.
    ins = wb.create_sheet(title="Instructions")
    instructions = [
        f"{project_type.name} bulk material requests",
        "",
        "Required columns (every row must have these):",
        "  - material   (must match an inventory item name on record)",
        "  - quantity   (numeric, > 0)",
        "  - region",
        "  - district",
        "  - community  (must match a community on record under this project)",
    ]
    if project_type.code == PROJECT_TYPE_SHEP:
        instructions.extend([
            "",
            "SHEP-specific:",
            "  - package_number  (optional in this template; if blank, looked up from the community)",
        ])
    elif project_type.code == PROJECT_TYPE_COST_SHARING:
        instructions.extend([
            "",
            "Cost Sharing-specific:",
            "  - beneficiary_contribution (optional; brief description of the cost-sharing arrangement)",
        ])
    elif project_type.code == PROJECT_TYPE_STREETLIGHTS:
        instructions.extend([
            "",
            "Streetlights-specific:",
            "  - pole_height_m  (numeric; metres)",
            "  - lumen_rating   (numeric)",
            "  - pole_type      (free text; e.g. galvanised steel, octagonal)",
        ])
    instructions.extend([
        "",
        "Optional columns:",
        "  - warehouse  (must match a warehouse name on record; blank = any)",
        "  - notes      (free text appended to the request notes)",
        "",
        "Signatory columns (batch-level — fill on the FIRST row only):",
        "  - memo_signatory_title    (must match a Signatory title flagged for the approval memo)",
        "  - letter_signatory_title  (must match a Signatory title flagged for the release letter)",
        "  Leave blank to use the active default set in the Signatory admin.",
        "  Unknown titles do NOT block the upload — they're surfaced as a warning",
        "  and you'll be asked to confirm a designation on the Release Letter page.",
        "",
        "On upload:",
        "  - Project type is set automatically from which template you downloaded.",
        "  - Consignee is auto-resolved from the community + project type.",
        "  - Project-specific fields are appended to the notes field for now.",
        "  - Signatory titles are stashed on the batch and pre-select the designation",
        "    pickers on the Release Letter generation page.",
        "  - Failed rows are returned as a downloadable error CSV with row number + column + reason.",
    ])

    for idx, line in enumerate(instructions, 1):
        cell = ins.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = Font(bold=True)
    ins.column_dimensions['A'].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename="material_request_template_{project_type.code}.xlsx"'
    )
    return response


def upload_requests(request):
    """
    Project-aware bulk material-request upload.

    POST: file (Excel), project (project type code).
    Validates each row, looks up community by name within the chosen project,
    auto-resolves consignee, creates MaterialOrder rows in a single
    transaction. Errors return as a downloadable CSV.
    """
    if request.method != 'POST':
        return redirect('request_material')

    uploaded_file = request.FILES.get('file')
    project_code = (request.POST.get('project') or '').strip().lower()

    if not uploaded_file or not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "Please upload an Excel file (.xlsx or .xls).")
        return redirect('request_material')

    if not project_code:
        messages.error(request, "Pick a project type before uploading.")
        return redirect('request_material')

    try:
        project_type = ProjectType.objects.get(code=project_code, active=True)
    except ProjectType.DoesNotExist:
        messages.error(request, f"Unknown or inactive project type: '{project_code}'.")
        return redirect('request_material')

    try:
        import pandas as pd
        df = pd.read_excel(uploaded_file)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f"Could not read Excel file: {exc}")
        return redirect('request_material')

    missing = require_columns(df, ['material', 'quantity', 'region', 'district', 'community'])
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}.")
        return redirect('request_material')

    df.columns = [str(c).strip().lower() for c in df.columns]
    result = BulkImportResult(total_rows=len(df))

    # Pre-fetch lookup tables.
    items_by_name = {it.name.strip().lower(): it for it in InventoryItem.objects.all()}
    warehouses_by_name = {w.name.strip().lower(): w for w in Warehouse.objects.all()}

    # Batch-level signatory hint — read the first non-empty value in each
    # signatory column. Resolved against the active Signatory roster,
    # filtered to officers eligible for the relevant document type. The
    # resolved pk is stashed on the first saved order's notes as a
    # structured marker that the Release Letter generation step reads to
    # pre-select the designation pickers.
    def _first_nonempty(col):
        if col not in df.columns:
            return ''
        for v in df[col]:
            s = normalize_cell(v)
            if s:
                return s
        return ''

    memo_sig_title   = _first_nonempty('memo_signatory_title')
    letter_sig_title = _first_nonempty('letter_signatory_title')

    def _resolve_sig(title, role_flag):
        if not title:
            return None, None
        sig = Signatory.objects.filter(
            active=True, title__iexact=title, **{role_flag: True},
        ).order_by('-updated_at').first()
        return sig, title

    memo_sig,   memo_sig_raw   = _resolve_sig(memo_sig_title,   'is_default_for_release_memo')
    letter_sig, letter_sig_raw = _resolve_sig(letter_sig_title, 'is_default_for_release_letter')

    rows_to_save = []
    request_code = generate_request_code()  # one batch code shared across rows

    for idx, row in df.iterrows():
        excel_row = idx + 2
        material_name = normalize_cell(row.get('material'))
        quantity_raw = normalize_cell(row.get('quantity'))
        region = normalize_cell(row.get('region'))
        district = normalize_cell(row.get('district'))
        community_name = normalize_cell(row.get('community'))
        warehouse_name = normalize_cell(row.get('warehouse')) if 'warehouse' in df.columns else ''
        notes = normalize_cell(row.get('notes')) if 'notes' in df.columns else ''

        if not (material_name or quantity_raw or community_name):
            continue  # silently skip empty rows

        if not material_name:
            result.add_error(excel_row, 'material', 'Required.')
        if not quantity_raw:
            result.add_error(excel_row, 'quantity', 'Required.')
        if not region:
            result.add_error(excel_row, 'region', 'Required.')
        if not district:
            result.add_error(excel_row, 'district', 'Required.')
        if not community_name:
            result.add_error(excel_row, 'community', 'Required.')

        # Material lookup.
        item = items_by_name.get(material_name.lower()) if material_name else None
        if material_name and item is None:
            result.add_error(excel_row, 'material', f"No inventory item named '{material_name}' on record.", material_name)

        # Quantity parse.
        try:
            qty = float(quantity_raw) if quantity_raw else 0
            if qty <= 0:
                result.add_error(excel_row, 'quantity', 'Must be > 0.', quantity_raw)
        except (TypeError, ValueError):
            qty = 0
            result.add_error(excel_row, 'quantity', 'Not a number.', quantity_raw)

        # Community lookup within this project.
        community = None
        if region and district and community_name:
            community = Community.objects.filter(
                project_type=project_type, region__iexact=region,
                district__iexact=district, community__iexact=community_name,
                is_active=True,
            ).first()
            if community is None:
                result.add_error(
                    excel_row, 'community',
                    f"No active {project_type.name} community '{community_name}' "
                    f"in {district}, {region}. Add it via Community management first.",
                    community_name,
                )

        warehouse = warehouses_by_name.get(warehouse_name.lower()) if warehouse_name else None
        if warehouse_name and warehouse is None:
            result.add_error(excel_row, 'warehouse', f"No warehouse named '{warehouse_name}' on record.", warehouse_name)

        # If row has any errors, skip it.
        if any(e.row_number == excel_row for e in result.errors):
            continue

        # Resolve consignee.
        resolved = resolve_consignee(project_type, community=community)

        # Project-specific extras → notes.
        extra_notes_parts = []
        if project_type.code == PROJECT_TYPE_SHEP:
            pkg = normalize_cell(row.get('package_number')) if 'package_number' in df.columns else ''
            if not pkg and community and community.package_number:
                pkg = community.package_number
            if pkg:
                extra_notes_parts.append(f"[SHEP] Package: {pkg}")
        elif project_type.code == PROJECT_TYPE_COST_SHARING:
            contrib = normalize_cell(row.get('beneficiary_contribution')) if 'beneficiary_contribution' in df.columns else ''
            if contrib:
                extra_notes_parts.append(f"[Cost Sharing] Beneficiary contribution: {contrib}")
        elif project_type.code == PROJECT_TYPE_STREETLIGHTS:
            ph = normalize_cell(row.get('pole_height_m')) if 'pole_height_m' in df.columns else ''
            lr = normalize_cell(row.get('lumen_rating')) if 'lumen_rating' in df.columns else ''
            pt = normalize_cell(row.get('pole_type')) if 'pole_type' in df.columns else ''
            sub = []
            if ph: sub.append(f"Pole height: {ph}m")
            if lr: sub.append(f"Lumen rating: {lr}")
            if pt: sub.append(f"Pole type: {pt}")
            if sub:
                extra_notes_parts.append(f"[Streetlights] {', '.join(sub)}")

        full_notes = notes
        if extra_notes_parts:
            extras = '\n'.join(extra_notes_parts)
            full_notes = f"{notes}\n\n{extras}".strip() if notes else extras

        # Resolve unit — the FK is NOT NULL on MaterialOrder. Fall back to the
        # first Unit if this inventory item has none (possible on older SQLite
        # rows) so one unitless item doesn't abort the whole batch.
        resolved_unit = item.unit
        if resolved_unit is None:
            from Inventory.models import Unit as _Unit
            resolved_unit = _Unit.objects.order_by('pk').first()
            if resolved_unit is None:
                result.add_error(
                    excel_row, 'material',
                    'Inventory item has no unit and no Unit records exist. '
                    'Add a Unit in admin, then re-upload.',
                    material_name,
                )
                continue

        # Build the MaterialOrder. Carry the item's name/code/category/unit so
        # the NOT NULL columns are populated — previously only `name` (set to
        # the item object) was passed, which crashed on unit_id.
        mo = MaterialOrder(
            name=item.name,
            code=item.code,
            category=item.category,
            unit=resolved_unit,
            quantity=qty,
            processed_quantity=0,
            remaining_quantity=qty,
            project_type=project_type_to_charfield(project_type),
            region=region,
            district=district,
            community=community_name,
            warehouse=warehouse,
            notes=full_notes,
            request_code=request_code,
            user=request.user,
            created_by=request.user,
            date_requested=timezone.now(),
        )
        # Carry the package number for EVERY project type (not just SHEP) so
        # non-SHEP releases are package-trackable and reconcile to BoQ. Prefer
        # an explicit package_number column, then the community's package.
        row_pkg = normalize_cell(row.get('package_number')) if 'package_number' in df.columns else ''
        resolved_pkg = row_pkg or (community.package_number if community else '') or ''
        if resolved_pkg:
            mo.package_number = resolved_pkg
        if resolved.kind == 'consultant':
            mo.consultant = resolved.name
        elif resolved.kind == 'mp':
            mo.contractor = resolved.render()

        rows_to_save.append(mo)

    # Attach the batch-level signatory hint to the FIRST order so the
    # downstream release-letter generator can pre-select the pickers.
    # Format is grep-friendly so the parser stays simple:
    #   [SignatoryHint] memo=<pk-or-RAW:title>; letter=<pk-or-RAW:title>
    if rows_to_save and (memo_sig_raw or letter_sig_raw):
        memo_token   = (str(memo_sig.pk)   if memo_sig   else (f"RAW:{memo_sig_raw}"   if memo_sig_raw   else ''))
        letter_token = (str(letter_sig.pk) if letter_sig else (f"RAW:{letter_sig_raw}" if letter_sig_raw else ''))
        marker = f"[SignatoryHint] memo={memo_token}; letter={letter_token}"
        first = rows_to_save[0]
        first.notes = (f"{marker}\n\n{first.notes}".strip() if first.notes else marker)

    if rows_to_save:
        try:
            with transaction.atomic():
                # request_code is unique per order, so the shared batch code
                # gets a per-row suffix (REQ-…-1, REQ-…-2, …). The common base
                # still lets the Release Letter wizard group the batch by prefix.
                for n, mo in enumerate(rows_to_save, 1):
                    mo.request_code = f"{request_code}-{n}"
                    mo.save()
            result.created_count = len(rows_to_save)
        except Exception as exc:  # noqa: BLE001
            messages.error(request, f"Bulk request upload failed: {exc}")
            return redirect('request_material')

    # Advisory messages when a signatory title was supplied but didn't
    # match an eligible officer in the roster.
    if memo_sig_raw and not memo_sig:
        messages.warning(
            request,
            f"Memo signatory title '{memo_sig_raw}' is not in the active roster as a memo signer. "
            "The Release Letter page will fall back to the active default — you can pick a different "
            "designation there.",
        )
    if letter_sig_raw and not letter_sig:
        messages.warning(
            request,
            f"Letter signatory title '{letter_sig_raw}' is not in the active roster as a letter signer. "
            "The Release Letter page will fall back to the active default — you can pick a different "
            "designation there.",
        )

    if result.has_errors:
        response = HttpResponse(result.errors_as_csv(), content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="request_upload_errors_{project_type.code}.csv"'
        )
        messages.warning(
            request,
            f"Bulk request upload completed with errors: {result.summary()} "
            "Error CSV downloaded — fix the rows and re-upload only those.",
        )
        return response

    if result.created_count or result.skipped_count:
        messages.success(request, f"Bulk request upload: {result.summary()}")
    else:
        messages.warning(request, "Bulk request upload: no rows processed (file may be empty).")
    return redirect('material_orders')


def resolve_consignee_for_community(request):
    """
    AJAX endpoint used by the Step 2 form to live-render the consignee
    preview block. Returns kind / display_label / display_text / reason.

    Query params:
      - project: ProjectType.code
      - community: Community.id
    """
    project_code = (request.GET.get('project') or '').strip().lower()
    community_id = request.GET.get('community')

    if not project_code:
        return JsonResponse({'error': 'project parameter required'}, status=400)
    try:
        project_type = ProjectType.objects.get(code=project_code, active=True)
    except ProjectType.DoesNotExist:
        return JsonResponse({'error': f"unknown project type '{project_code}'"}, status=404)

    community = None
    if community_id:
        try:
            community = Community.objects.get(pk=int(community_id))
        except (ValueError, Community.DoesNotExist):
            community = None

    resolved = resolve_consignee(project_type, community=community)

    return JsonResponse({
        'kind': resolved.kind,
        'is_resolved': resolved.is_resolved,
        'display_label': resolved.display_label,
        'name': resolved.name,
        'detail': resolved.detail,
        'rendered': resolved.render(),
        'reason': resolved.reason,
    })
