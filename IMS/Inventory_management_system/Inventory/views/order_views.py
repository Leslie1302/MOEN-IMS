import json
import logging
import uuid
from decimal import Decimal, InvalidOperation

import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Count, Sum, F
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import View, ListView
from django.http import JsonResponse, Http404
from django.forms import formset_factory
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.exceptions import ValidationError

from Inventory.models import (
    InventoryItem, MaterialOrder, MaterialOrderAudit,
    ReleaseLetter, Warehouse, MaterialTransport, SiteReceipt, BillOfQuantity
)
from Inventory.forms import (
    MaterialOrderForm, BulkMaterialRequestForm, MaterialReceiptFormSet
)

# Define MaterialOrderFormSet locally if not imported
MaterialOrderFormSet = formset_factory(MaterialOrderForm, extra=1)

# Configure logger
logger = logging.getLogger(__name__)


# ===== API ENDPOINTS =====

@login_required
def get_inventory_item_details(request, item_id):
    """API endpoint to get inventory item details for autofill"""
    try:
        item = InventoryItem.objects.get(id=item_id)
        return JsonResponse({
            'success': True,
            'data': {
                'id': item.id,
                'name': item.name,
                'category': item.category.name if item.category else '',
                'code': item.code,
                'unit': item.unit.name if item.unit else '',
                'quantity': item.quantity,
                'warehouse': item.warehouse.name if item.warehouse else ''
            }
        })
    except InventoryItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def generate_request_code():
    """Generate a unique request code in the format REQ-YYYYMMDD-XXXXXX"""
    date_str = timezone.now().strftime('%Y%m%d')
    unique_id = str(uuid.uuid4().int)[:6].upper()
    return f"REQ-{date_str}-{unique_id}"


def _batch_base(code):
    """Batch base of a request code. Bulk uploads suffix each row
    (REQ-…-1, -2, …); single requests have no suffix. Strip a trailing numeric
    segment ONLY when there are more than 3 segments so single codes (and the
    numeric date/random parts) are left intact."""
    parts = (code or '').split('-')
    if len(parts) > 3 and parts[-1].isdigit():
        return '-'.join(parts[:-1])
    return code or ''


def annotate_bulk_batches(orders):
    """Tag Release orders that belong to a bulk upload batch so the orders page
    can offer ONE batch action and grey out the siblings.

    Sets on each order: ``is_bulk``, ``bulk_size``, ``bulk_base`` and
    ``batch_release_letter`` (the single letter the batch is released under,
    once generated — every row links to it).
    """
    from collections import defaultdict
    orders = list(orders)
    by_base = defaultdict(list)
    for o in orders:
        if getattr(o, 'request_type', None) == 'Release':
            by_base[_batch_base(o.request_code)].append(o)
    bulk_bases = [b for b, lst in by_base.items() if len(lst) > 1]
    rl_by_base = {}
    if bulk_bases:
        for rl in ReleaseLetter.objects.filter(request_code__in=bulk_bases):
            rl_by_base.setdefault(rl.request_code, rl)
    for b, lst in by_base.items():
        size = len(lst)
        rl = rl_by_base.get(b)
        for o in lst:
            o.is_bulk = size > 1
            o.bulk_size = size
            o.bulk_base = b
            o.batch_release_letter = rl
    return orders


class RequestMaterialView(LoginRequiredMixin, View):
    template_name = 'Inventory/request_material.html'

    def get(self, request):
        # Filter items based on user permissions
        if request.user.is_superuser:
            items = InventoryItem.objects.all()
        else:
            items = InventoryItem.objects.filter(group__in=request.user.groups.all())

        formset = MaterialOrderFormSet(form_kwargs={'user': request.user})
        bulk_form = BulkMaterialRequestForm()
        inventory_items = list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))

        # Non-superusers default to bulk tab, superusers default to single tab
        default_tab = 'single' if request.user.is_superuser else 'bulk'

        return render(request, self.template_name, {
            'formset': formset,
            'bulk_form': bulk_form,
            'items': items,
            'inventory_items': json.dumps(inventory_items),
            'active_tab': default_tab
        })

    def post(self, request):
        # Check which form was submitted
        if 'bulk_submit' in request.POST:
            return self.handle_bulk_request(request)
        else:
            return self.handle_single_request(request)

    def handle_single_request(self, request):
        formset = MaterialOrderFormSet(request.POST, request.FILES, form_kwargs={'user': request.user})
        if formset.is_valid():
            request_code = generate_request_code()
            created_orders = []
            with transaction.atomic():
                for form in formset:
                    if form.cleaned_data:
                        material_order = form.save(commit=False)
                        selected_item = form.cleaned_data['name']  # This is an InventoryItem object
                        selected_warehouse = form.cleaned_data.get('warehouse')
                        
                        # Look up the specific inventory item by name and warehouse
                        if selected_item and selected_warehouse:
                            try:
                                inventory_item = InventoryItem.objects.get(
                                    name=selected_item.name,
                                    warehouse=selected_warehouse
                                )
                                material_order.name = inventory_item.name
                                material_order.category = inventory_item.category
                                material_order.code = inventory_item.code
                                material_order.unit = inventory_item.unit
                            except InventoryItem.DoesNotExist:
                                # Fallback to selected item if specific warehouse combo doesn't exist
                                material_order.name = selected_item.name
                                material_order.category = selected_item.category
                                material_order.code = selected_item.code
                                material_order.unit = selected_item.unit
                        elif selected_item:
                            material_order.name = selected_item.name
                            material_order.category = selected_item.category
                            material_order.code = selected_item.code
                            material_order.unit = selected_item.unit
                        
                        material_order.user = request.user
                        material_order.group = request.user.groups.first() if request.user.groups.exists() else None
                        material_order.request_type = 'Release'
                        material_order.request_code = request_code
                        # Ensure newly created requests start as Draft
                        material_order.status = 'Draft'
                        # Initialize quantities so remaining is not zero
                        material_order.processed_quantity = 0
                        material_order.remaining_quantity = material_order.quantity
                        
                        # Set current user for release letter creation
                        material_order._current_user = request.user
                        
                        # Save the material order first to get the ID and proper request_code
                        material_order.save()
                        created_orders.append(material_order)
                        
                        # Now handle release letter creation if file was uploaded
                        if form.cleaned_data.get('release_letter_pdf'):
                            title = form.cleaned_data.get('release_letter_title') or f"Release Letter for {material_order.name}"
                            auth_qty = form.cleaned_data.get('release_letter_quantity') or material_order.quantity
                            material_type = form.cleaned_data.get('release_letter_material_type') or 'Other'
                            phase = form.cleaned_data.get('release_letter_project_phase')
                            
                            release_letter = ReleaseLetter.objects.create(
                                request_code=material_order.request_code,
                                title=title,
                                total_quantity=auth_qty,
                                material_type=material_type,
                                project_phase=phase,
                                pdf_file=form.cleaned_data['release_letter_pdf'],
                                upload_time=timezone.now(),
                                uploaded_by=request.user,
                                notes=f"Uploaded with material request {material_order.request_code}"
                            )
                            
                            # Link the release letter to the material order
                            material_order.release_letter = release_letter
                            material_order.save()
                            
            # Warn if any request targets a package that is not in any BoQ.
            off_boq_packages = sorted({
                o.package_number for o in created_orders
                if o.package_number
                and not BillOfQuantity.objects.filter(
                    package_number=o.package_number).exists()
            })
            if off_boq_packages:
                messages.warning(
                    request,
                    "Heads-up: package number(s) "
                    + ", ".join(off_boq_packages)
                    + " are not in any Bill of Quantity. These releases will "
                    "not draw down a contract when delivered. Check the "
                    "BoQ overissuance report for off-BoQ deliveries."
                )

            messages.success(request, "Material requests submitted successfully!")
            return redirect('material_orders')
        else:
            print("Formset errors:", formset.errors)
            messages.error(request, "There was an error with your submission.")

        # Prepare context for re-rendering the form with errors
        # Show all inventory items to all users for transparency
        items = InventoryItem.objects.all()

        return render(request, self.template_name, {
            'formset': formset,
            'bulk_form': BulkMaterialRequestForm(),
            'items': items,
            'inventory_items': json.dumps(list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))),
            'active_tab': 'single'
        })

    def handle_bulk_request(self, request):
        logger = logging.getLogger(__name__)
        logger.info("Starting bulk request processing...")
        
        bulk_form = BulkMaterialRequestForm(request.POST, request.FILES)
        if not bulk_form.is_valid():
            logger.error(f"Bulk form validation failed: {bulk_form.errors}")
            for field, errors in bulk_form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return self._render_request_form(request, bulk_form=bulk_form)
        
        # Initialize variables outside the transaction
        success_count = 0
        error_messages = []
        
        try:
            df = bulk_form.cleaned_data['df']
            request_type = bulk_form.cleaned_data['request_type']
            priority = bulk_form.cleaned_data['priority']  # Get priority from form
            release_letter_pdf = bulk_form.cleaned_data.get('release_letter_pdf')
            release_letter_title = bulk_form.cleaned_data.get('release_letter_title')
            
            # Inform user if rows were filtered out
            filtered_count = bulk_form.cleaned_data.get('filtered_rows', 0)
            if filtered_count > 0:
                messages.info(request, f"Note: {filtered_count} row(s) with zero or negative quantities were automatically skipped.")
            
            logger.info(f"Processing bulk request with {len(df)} rows")
            
            # Generate a base request code for reference
            base_request_code = generate_request_code()
            logger.info(f"Base request code for this batch: {base_request_code}")
            
            # Add a request code column to the DataFrame
            df['request_code'] = [f"{base_request_code}-{i+1}" for i in range(len(df))]
            
            # Create release letter if PDF is uploaded
            release_letter = None
            if release_letter_pdf:
                try:
                    total_batch_quantity = bulk_form.cleaned_data.get('release_letter_quantity') or df['quantity'].sum()
                    material_type = bulk_form.cleaned_data.get('release_letter_material_type') or 'Other'
                    phase = bulk_form.cleaned_data.get('release_letter_project_phase')

                    # Carry through any per-event letterhead overrides.
                    # Resolution order, strongest first:
                    #   1) explicit form POST (the user picked from the
                    #      designation dropdown on bulk_request.html)
                    #   2) Excel columns memo_signatory_title /
                    #      letter_signatory_title — first non-empty row,
                    #      matched against active signatories flagged
                    #      eligible for that document type
                    #   3) the active default on the Signatory admin
                    from Inventory.models import Signatory
                    memo_to       = (request.POST.get('memo_to_override') or '').strip()
                    memo_from     = (request.POST.get('memo_from_override') or '').strip()
                    memo_sig_id   = (request.POST.get('memo_signatory_override') or '').strip()
                    letter_sig_id = (request.POST.get('letter_signatory_override') or '').strip()
                    memo_signatory   = Signatory.objects.filter(pk=memo_sig_id).first() if memo_sig_id else None
                    letter_signatory = Signatory.objects.filter(pk=letter_sig_id).first() if letter_sig_id else None

                    def _first_title(col):
                        if col not in df.columns:
                            return ''
                        for v in df[col]:
                            try:
                                s = str(v).strip()
                            except Exception:
                                s = ''
                            if s and s.lower() != 'nan':
                                return s
                        return ''

                    if memo_signatory is None:
                        t = _first_title('memo_signatory_title')
                        if t:
                            memo_signatory = Signatory.objects.filter(
                                active=True, title__iexact=t,
                                is_default_for_release_memo=True,
                            ).order_by('-updated_at').first()
                            if memo_signatory is None:
                                messages.warning(
                                    request,
                                    f"Memo signatory title '{t}' from the Excel isn't in the active "
                                    "roster as a memo signer — falling back to the default.",
                                )
                    if letter_signatory is None:
                        t = _first_title('letter_signatory_title')
                        if t:
                            letter_signatory = Signatory.objects.filter(
                                active=True, title__iexact=t,
                                is_default_for_release_letter=True,
                            ).order_by('-updated_at').first()
                            if letter_signatory is None:
                                messages.warning(
                                    request,
                                    f"Letter signatory title '{t}' from the Excel isn't in the active "
                                    "roster as a letter signer — falling back to the default.",
                                )

                    release_letter = ReleaseLetter.objects.create(
                        title=release_letter_title or f"Release Letter - {base_request_code}",
                        total_quantity=Decimal(str(total_batch_quantity)),
                        material_type=material_type,
                        project_phase=phase,
                        pdf_file=release_letter_pdf,
                        uploaded_by=request.user,
                        request_code=base_request_code,
                        memo_to_override=memo_to,
                        memo_from_override=memo_from,
                        memo_signatory_override=memo_signatory,
                        letter_signatory_override=letter_signatory,
                    )
                    logger.info(f"Created release letter ID {release_letter.id} for request code {base_request_code}")
                except Exception as e:
                    error_msg = f"Error creating release letter: {str(e)}"
                    messages.error(request, error_msg)
                    logger.error(error_msg, exc_info=True)
                    return self._render_request_form(request, bulk_form=bulk_form)
            
            # Process each row in the Excel file
            logger.info(f"Starting to process {len(df)} rows from Excel")
            
            for idx, row in df.iterrows():
                row_dict = row.to_dict()
                
                try:
                    # Skip empty rows
                    if not row.get('name'):
                        continue
                        
                    # Find the inventory item
                    item_name = str(row['name']).strip()
                    item = self._find_inventory_item(item_name, request.user)
                    
                    if not item:
                        error_msg = f"Item not found or not accessible: {item_name}"
                        error_messages.append(error_msg)
                        continue
                        
                    # Handle group assignment
                    group = self._get_order_group(request.user, item, item_name)
                    
                    # Handle warehouse lookup
                    warehouse = None
                    warehouse_name = row.get('warehouse')
                    if warehouse_name and pd.notna(warehouse_name):
                        warehouse_name_str = str(warehouse_name).strip()
                        try:
                            warehouse = Warehouse.objects.filter(name__iexact=warehouse_name_str).first()
                        except Exception as wh_error:
                            logger.error(f"Error looking up warehouse: {str(wh_error)}")
                    
                    # Resolve unit — the FK is NOT NULL on MaterialOrder.
                    # InventoryItem.unit_id can be NULL on SQLite databases where
                    # an early migration added the column without enforcing NOT NULL.
                    # Fall back to the first Unit in the system, or skip the row
                    # with a clear error so one bad item doesn't abort the whole upload.
                    resolved_unit = item.unit
                    if resolved_unit is None:
                        from Inventory.models import Unit as _Unit
                        resolved_unit = _Unit.objects.order_by('pk').first()
                        if resolved_unit is None:
                            error_messages.append(
                                f"❌ SKIPPED '{item_name}': inventory item has no unit assigned "
                                f"and no Unit records exist in the system. "
                                f"Add a Unit in admin, then re-upload."
                            )
                            continue
                        logger.warning(
                            f"InventoryItem pk={item.pk} ('{item.name}') has unit_id=NULL — "
                            f"falling back to Unit pk={resolved_unit.pk} ('{resolved_unit}') "
                            f"for this bulk order. Fix the InventoryItem in admin."
                        )

                    # Create the order in a new transaction for each item
                    try:
                        with transaction.atomic():
                            order_data = {
                                'name': item.name,
                                'quantity': row['quantity'],
                                'category': item.category,
                                'code': item.code,
                                'unit': resolved_unit,
                                'user': request.user,
                                'group': group,
                                'warehouse': warehouse,
                                'request_type': request_type,
                                'request_code': row['request_code'],  # Use the unique request code from the DataFrame
                                'priority': priority,  # Use priority from form (applies to all items)
                                'region': row.get('region', ''),
                                'district': row.get('district', ''),
                                'community': row.get('community', ''),
                                'consultant': row.get('consultant', ''),
                                'contractor': row.get('contractor', ''),
                                'package_number': row.get('package_number', ''),
                                'last_updated_by': request.user,
                                # Ensure bulk-created requests start as Draft
                                'status': 'Draft',
                                # Initialize quantities so remaining is not zero
                                'processed_quantity': 0,
                                'remaining_quantity': row['quantity']
                            }
                            
                            # Create the order
                            order = MaterialOrder.objects.create(**order_data)
                            
                            # Associate with release letter if available
                            if release_letter:
                                order.release_letter = release_letter
                                order.save(update_fields=['release_letter'])
                            
                            success_count += 1
                            
                    except Exception as e:
                        error_msg = f"❌ ERROR saving order for {item_name}: {str(e)}"
                        error_messages.append(error_msg)
                        logger.error(error_msg, exc_info=True)
                        continue
                        
                except Exception as e:
                    error_msg = f"❌ ERROR processing row for {row.get('name', 'unknown')}: {str(e)}"
                    error_messages.append(error_msg)
                    logger.error(error_msg, exc_info=True)
                    continue
            
            # Show success/error messages
            if success_count > 0:
                msg = f"Successfully created {success_count} material request(s) with unique request codes starting with {base_request_code}"
                messages.success(request, msg)
                
                # Add an info message about how to track related requests
                if success_count > 1:
                    messages.info(request, 
                        "Each item in the bulk upload has been assigned a unique request code. "
                        "You can find related requests by searching for the base code."
                    )
                
                # Only redirect if we had any successful saves
                return redirect('material_orders')
                
            # If we got here, there were no successful saves
            if error_messages:
                for error in error_messages[:5]:  # Show first 5 errors to avoid flooding
                    messages.error(request, error)
                if len(error_messages) > 5:
                    messages.warning(request, f"... and {len(error_messages) - 5} more errors occurred.")
                    
            return self._render_request_form(request, bulk_form=bulk_form)
                    
        except Exception as e:
            error_msg = f"Unexpected error processing bulk request: {str(e)}"
            messages.error(request, error_msg)
            logger.error(error_msg, exc_info=True)
            return self._render_request_form(request, bulk_form=bulk_form)
            
        return self._render_request_form(request, bulk_form=bulk_form)
        
    def _find_inventory_item(self, item_name, user):
        """Helper method to find an inventory item by name with proper permissions"""
        logger = logging.getLogger(__name__)
        try:
            # First try exact match
            if user.is_superuser:
                item = InventoryItem.objects.filter(name__iexact=item_name).first()
            else:
                item = InventoryItem.objects.filter(
                    name__iexact=item_name,
                    group__in=user.groups.all()
                ).first()
                
            if item:
                return item
                
            # If no exact match, try case-insensitive contains
            if user.is_superuser:
                item = InventoryItem.objects.filter(name__icontains=item_name).first()
            else:
                item = InventoryItem.objects.filter(
                    name__icontains=item_name,
                    group__in=user.groups.all()
                ).first()
                
            if item:
                return item
                
            return None
            
        except Exception as e:
            logger.error(f"Error finding inventory item {item_name}: {str(e)}", exc_info=True)
            return None

    def _get_order_group(self, user, item, item_name):
        """Helper method to determine the appropriate group for an order"""
        # First try to get the group from the item
        if item.group:
            return item.group
            
        # If item has no group, try to find a matching group from the user's groups
        # that has the same name as the item's category
        if item.category and user.groups.exists():
            matching_group = user.groups.filter(name__iexact=item.category.name).first()
            if matching_group:
                return matching_group
                
        # Default to the user's first group if available
        if user.groups.exists():
            return user.groups.first()
            
        return None

    def _render_request_form(self, request, bulk_form=None):
        """Helper method to render the request form with the current context"""
        if request.user.is_superuser:
            items = InventoryItem.objects.all()
        else:
            items = InventoryItem.objects.filter(group__in=request.user.groups.all())

        from Inventory.models import Signatory
        context = {
            'formset': MaterialOrderFormSet(form_kwargs={'user': request.user}),
            'bulk_form': bulk_form or BulkMaterialRequestForm(),
            'items': items,
            'inventory_items': json.dumps(list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))),
            'active_tab': 'bulk' if bulk_form else 'single',
            # Designation-led pickers for the letterhead override section.
            # Each list is filtered to officers flagged eligible for that
            # document type and sorted by title so the user picks a role,
            # not a name.
            'memo_signatories': Signatory.objects.filter(
                active=True, is_default_for_release_memo=True,
            ).order_by('title'),
            'letter_signatories': Signatory.objects.filter(
                active=True, is_default_for_release_letter=True,
            ).order_by('title'),
            # Kept for any legacy template that still loops `signatories`.
            'signatories': Signatory.objects.filter(active=True).order_by('title'),
        }
        return render(request, self.template_name, context)


# Phase L -- split material orders into active vs archived.
# Active = anything still in motion; Archived = terminal states.
ACTIVE_ORDER_STATUSES = (
    'Draft', 'Pending', 'Approved', 'In Progress',
    'Partially Fulfilled', 'Ready for Pickup', 'In Transit',
)
ARCHIVED_ORDER_STATUSES = (
    'Delivered', 'Completed', 'Rejected', 'Cancelled',
)


class MaterialOrdersView(LoginRequiredMixin, ListView):
    """
    Active material orders. Terminal-state orders (Completed / Cancelled /
    Rejected / Delivered) live on /material-orders/archive/ instead so this
    page stays focused on work in motion.
    """
    template_name = 'Inventory/material_orders.html'
    context_object_name = 'orders'
    paginate_by = 50
    paginate_orphans = 5
    allow_empty = True
    is_archive = False

    def get_queryset(self):
        try:
            qs = MaterialOrder.objects.select_related(
                'user', 'unit', 'category', 'warehouse'
            ).order_by('-date_requested')
            if self.is_archive:
                qs = qs.filter(status__in=ARCHIVED_ORDER_STATUSES)
            else:
                qs = qs.filter(status__in=ACTIVE_ORDER_STATUSES)
            return qs
        except Exception as e:
            logger.error(f"Error in MaterialOrdersView: {str(e)}", exc_info=True)
            return MaterialOrder.objects.none()

    def paginate_queryset(self, queryset, page_size):
        """Override to handle invalid page numbers gracefully"""
        from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
        
        paginator = Paginator(queryset, page_size, orphans=self.paginate_orphans)
        page_number = self.request.GET.get('page', 1)
        
        try:
            page = paginator.page(page_number)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page
            page = paginator.page(1)
        except EmptyPage:
            # If page is out of range, deliver last page
            page = paginator.page(paginator.num_pages)
        
        return (paginator, page, page.object_list, page.has_other_pages())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        stats = MaterialOrder.objects.aggregate(
            total_orders=Count('id'),
            pending_orders=Count('id', filter=Q(status='Pending')),
            completed_orders=Count('id', filter=Q(status='Completed')),
            partial_orders=Count('id', filter=Q(status='Partially Fulfilled')),
            active_orders=Count('id', filter=Q(status__in=ACTIVE_ORDER_STATUSES)),
            archived_orders=Count('id', filter=Q(status__in=ARCHIVED_ORDER_STATUSES)),
        )
        context.update(stats)
        context['is_archive'] = self.is_archive
        context['active_url_name'] = 'material_orders'
        context['archive_url_name'] = 'material_orders_archive'
        annotate_bulk_batches(context.get('orders') or context.get('object_list') or [])
        return context


class MaterialOrdersArchiveView(MaterialOrdersView):
    """Archived material orders -- terminal states only."""
    is_archive = True


class MaterialOrdersOfficersView(LoginRequiredMixin, ListView):
    """
    Officers view of material orders. Renders the unified material_orders.html
    template. Honours the same active/archive split as MaterialOrdersView so
    completed/cancelled orders don't keep cluttering the day-to-day view.

    Previously this view also excluded Draft/Pending, which made the page
    blank for schedule officers who had just filed new requests (their own
    pending queue vanished until something was approved). Schedule officers
    land on this page from their dropdown, so the Pending queue stays.
    """
    template_name = 'Inventory/material_orders.html'
    context_object_name = 'orders'
    paginate_by = 50
    is_archive = False

    def get_queryset(self):
        user = self.request.user
        logger = logging.getLogger(__name__)

        try:
            queryset = MaterialOrder.objects.select_related(
                'user', 'unit', 'category', 'assigned_to', 'assigned_by'
            ).order_by('-date_requested')

            if self.is_archive:
                queryset = queryset.filter(status__in=ARCHIVED_ORDER_STATUSES)
            else:
                # Show every in-motion order, including freshly filed
                # Draft/Pending requests, so schedule officers see their own
                # work as soon as they submit it.
                queryset = queryset.filter(status__in=ACTIVE_ORDER_STATUSES)

            logger.info(f"User {user.username} accessing {queryset.count()} total orders")

            # Ensure remaining_quantity is calculated correctly
            for order in queryset:
                if order.remaining_quantity is None or order.remaining_quantity < 0:
                    order.remaining_quantity = max(0, order.quantity - (order.processed_quantity or 0))
                    order.save(update_fields=['remaining_quantity'])

            return queryset

        except Exception as e:
            logger.error(f"Error in MaterialOrdersOfficersView for user {user.username}: {str(e)}", exc_info=True)
            # Fallback to empty queryset to prevent crashes
            return MaterialOrder.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Active/archive stats so the toggle badges populate (the template
        # was previously rendering blank counts on the officers page).
        stats = MaterialOrder.objects.aggregate(
            total_orders=Count('id'),
            pending_orders=Count('id', filter=Q(status='Pending')),
            completed_orders=Count('id', filter=Q(status='Completed')),
            partial_orders=Count('id', filter=Q(status='Partially Fulfilled')),
            active_orders=Count('id', filter=Q(status__in=ACTIVE_ORDER_STATUSES)),
            archived_orders=Count('id', filter=Q(status__in=ARCHIVED_ORDER_STATUSES)),
        )
        context.update(stats)
        context['is_archive'] = self.is_archive
        # Tell the template to point the Active/Archive toggle at the
        # officers' URL pair, not the global stores URL pair. Without this
        # the Active button would whisk a schedule officer away to the
        # stores page and they'd think the officers page was broken.
        context['active_url_name'] = 'material_orders_officers'
        context['archive_url_name'] = 'material_orders_officers_archive'
        annotate_bulk_batches(context.get('orders') or context.get('object_list') or [])
        return context


class MaterialOrdersOfficersArchiveView(MaterialOrdersOfficersView):
    """Archived material orders for the officers view."""
    is_archive = True


class UpdateMaterialStatusView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Handle material order status updates and fulfillment processing.
    Supports: Seen, Approved, Rejected, Partial, Full status updates.
    """

    def test_func(self):
        """Ensure only staff/superusers or explicitly assigned users can update status."""
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return True
        return user.groups.exists() 
    
    def post(self, request, order_id, new_status):
        logger = logging.getLogger(__name__)
        
        # Validate status
        allowed_statuses = ["Seen", "Approved", "Rejected", "Partially Fulfilled", "Full"]
        if new_status not in allowed_statuses:
            return JsonResponse({
                "success": False, 
                "error": f"Invalid status '{new_status}'. Allowed: {', '.join(allowed_statuses)}"
            }, status=400)

        try:
            with transaction.atomic():
                order = get_object_or_404(MaterialOrder, id=order_id)
                
                # Parse request data
                try:
                    data = json.loads(request.body.decode('utf-8'))
                except json.JSONDecodeError as e:
                    return JsonResponse({
                        "success": False, 
                        "error": "Invalid JSON in request body"
                    }, status=400)

                # Handle simple status changes
                if new_status in ["Seen", "Approved", "Rejected"]:
                    order.status = new_status
                
                # Handle quantity processing (Partial/Full)
                elif new_status in ["Partially Fulfilled", "Full"]:
                    
                    # SECURITY: Check if order is assigned to current user
                    if order.assigned_to and order.assigned_to != request.user:
                        return JsonResponse({
                            'success': False, 
                            'error': f'This order is assigned to {order.assigned_to.get_full_name() or order.assigned_to.username}. Only the assigned user can process this order.'
                        }, status=403)
                    
                    # Validate current status allowing quantity processing
                    required_status = 'Approved' if order.request_type == 'Release' else 'Seen'
                    
                    if order.status not in [required_status, 'Partially Fulfilled']:
                        return JsonResponse({
                            'success': False, 
                            'error': f'Order must be in "{required_status}" or "Partially Fulfilled" status before processing quantities. Current status: "{order.status}"'
                        }, status=400)

                    # Calculate quantity to process
                    if new_status == "Partially Fulfilled":
                        try:
                            partial_quantity = Decimal(str(data.get('partial_quantity', 0)))
                        except (ValueError, TypeError, InvalidOperation) as e:
                            return JsonResponse({
                                "success": False, 
                                "error": "Invalid partial_quantity value. Must be a valid number."
                            }, status=400)
                    else:  # Full
                        partial_quantity = order.remaining_quantity

                    # Validate quantity
                    if partial_quantity <= 0:
                        return JsonResponse({
                            "success": False, 
                            "error": "Quantity must be greater than zero"
                        }, status=400)
                    
                    if partial_quantity > order.remaining_quantity:
                        return JsonResponse({
                            'success': False, 
                            'error': f'Quantity {partial_quantity} exceeds remaining quantity {order.remaining_quantity}'
                        }, status=400)

                    # Update order quantities
                    order.processed_quantity = (order.processed_quantity or 0) + partial_quantity
                    order.remaining_quantity = max(0, order.quantity - order.processed_quantity)
                    
                    # Set processing tracking fields
                    order.processed_by = request.user
                    order.processed_at = timezone.now()
                    
                    # SIGNED LETTER GUARD: a Release cannot draw down stock
                    # until the signed scan is on file. This bars accidental
                    # processing of an order whose paperwork hasn't been
                    # countersigned yet.
                    if order.request_type == "Release":
                        rl = order.release_letter
                        if rl is None:
                            return JsonResponse({
                                'success': False,
                                'error': 'Cannot release: no release letter has been created for this order. Generate the release letter first.'
                            }, status=400)
                        if not rl.pdf_file:
                            return JsonResponse({
                                'success': False,
                                'error': f'Cannot release: signed copy of release letter {rl.code or rl.reference_number or ""} is not attached. Upload the signed scan before processing.'
                            }, status=400)

                    # Update inventory
                    try:
                        # Match by code and warehouse (unique_together constraint).
                        # When the request was filed with "any warehouse", pick
                        # the warehouse with enough stock so the deduction
                        # actually goes somewhere instead of blowing up on
                        # MultipleObjectsReturned.
                        inventory_item = None
                        if order.warehouse:
                            inventory_item = InventoryItem.objects.filter(
                                code=order.code,
                                warehouse=order.warehouse,
                            ).first()
                        else:
                            candidates = list(
                                InventoryItem.objects.filter(code=order.code)
                                .order_by('-quantity')
                            )
                            if order.request_type == "Release":
                                # Pick the warehouse that can satisfy the draw.
                                for cand in candidates:
                                    if (cand.quantity or 0) >= partial_quantity:
                                        inventory_item = cand
                                        break
                                if inventory_item is None and candidates:
                                    # Fall back to the largest holder so the
                                    # error message below reports a real number.
                                    inventory_item = candidates[0]
                            else:
                                # Receipts: deposit into the order's stored
                                # warehouse if any; else the first sibling.
                                inventory_item = candidates[0] if candidates else None
                        if inventory_item is None:
                            logger.warning(
                                f"Inventory item with code '{order.code}' not found "
                                f"(warehouse={order.warehouse}). Skipping inventory update."
                            )
                        else:
                            if order.request_type == "Release":
                                if inventory_item.quantity < partial_quantity:
                                    return JsonResponse({
                                        'success': False,
                                        'error': f'Insufficient inventory. Available: {inventory_item.quantity}, Requested: {partial_quantity}'
                                    }, status=400)
                                inventory_item.quantity -= partial_quantity
                            elif order.request_type == "Receipt":
                                inventory_item.quantity += partial_quantity
                            inventory_item.save()
                            # Track how much stock has already been moved so the
                            # post_save signal only deducts the delta (prevents
                            # double deduction when the view and signal both run).
                            order.stock_deducted_quantity = (order.stock_deducted_quantity or 0) + partial_quantity
                            # Stamp the warehouse used for the draw so audit
                            # trails reflect where the stock actually moved.
                            if not order.warehouse_id and inventory_item.warehouse_id:
                                order.warehouse = inventory_item.warehouse

                    except InventoryItem.DoesNotExist:
                        logger.warning(f"Inventory item with code '{order.code}' not found in warehouse '{order.warehouse}'. Skipping inventory update.")
                    except InventoryItem.MultipleObjectsReturned:
                        return JsonResponse({
                            'success': False,
                            'error': f'Multiple inventory items found with code "{order.code}". Please contact administrator to resolve duplicate items.'
                        }, status=500)

                    # Update order status based on remaining quantity
                    if order.remaining_quantity <= 0:
                        order.status = 'Completed'
                    else:
                        order.status = 'Partially Fulfilled'

                # Update audit fields
                order.last_updated_by = request.user
                order.save()
                order.refresh_from_db()

                # Prepare response data
                try:
                    status_html = render_to_string('Inventory/includes/status_cell.html', {'order': order})
                except Exception as e:
                    status_html = f'<span class="badge bg-secondary">{order.status}</span>'

                response_data = {
                    'success': True,
                    'new_status': order.get_status_display(),
                    'status_html': status_html.strip(),
                    'processed_quantity': float(order.processed_quantity or 0),
                    'remaining_quantity': float(order.remaining_quantity or 0),
                    'is_completed': order.status in ['Completed', 'Rejected'] or order.remaining_quantity <= 0,
                    'last_updated_by': order.last_updated_by.username if order.last_updated_by else 'System',
                    'message': f'Order {order.request_code or order.id} status updated to {order.get_status_display()}'
                }

                logger.info(f"Successfully processed order {order_id}: {response_data}")
                return JsonResponse(response_data)

        except (MaterialOrder.DoesNotExist, Http404):
            return JsonResponse({
                "success": False, 
                "error": f"Material order with ID {order_id} not found"
            }, status=404)
        except ValidationError as e:
            return JsonResponse({
                "success": False, 
                "error": str(e.message) if hasattr(e, 'message') else str(e)
            }, status=400)
        except Exception as e:
            logger.error(f"Unexpected error updating material status for order {order_id}: {e}", exc_info=True)
            return JsonResponse({
                "success": False, 
                "error": "An unexpected server error occurred. Please try again."
            }, status=500)


@login_required
def update_material_receipt(request, order_id, new_status):
    """
    Wrapper for updating material receipt status.
    Reuses the logic from UpdateMaterialStatusView.
    """
    return UpdateMaterialStatusView.as_view()(request, order_id=order_id, new_status=new_status)


class MaterialReceiptView(LoginRequiredMixin, View):
    template_name = 'Inventory/receive_material.html'

    def get(self, request):
        # Show all inventory items to all users for transparency
        items = InventoryItem.objects.all()
        inventory_items = list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))
        
        formset = MaterialReceiptFormSet(form_kwargs={'user': request.user})
        bulk_form = BulkMaterialRequestForm()
        # Mocking or fetching receipt orders if needed for context
        orders = MaterialOrder.objects.filter(request_type='Receipt').order_by('-date_requested')

        return render(request, self.template_name, {
            'formset': formset,
            'bulk_form': bulk_form,
            'items': items,
            'inventory_items': json.dumps(inventory_items),
            'active_tab': 'single',
            'orders': orders,
        })

    def post(self, request):
        # Check which form was submitted
        if 'bulk_submit' in request.POST:
            return self.handle_bulk_receipt(request)
        else:
            return self.handle_single_receipt(request)

    def handle_single_receipt(self, request):
        formset = MaterialReceiptFormSet(request.POST, form_kwargs={'user': request.user})
        if formset.is_valid():
            for form in formset:
                if form.cleaned_data:
                    material_order = form.save(commit=False)
                    selected_item = form.cleaned_data['name']  # InventoryItem
                    selected_warehouse = form.cleaned_data.get('warehouse')

                    # Look up the specific inventory item by name and warehouse
                    if selected_item and selected_warehouse:
                        try:
                            inventory_item = InventoryItem.objects.get(
                                name=selected_item.name,
                                warehouse=selected_warehouse
                            )
                            material_order.name = inventory_item.name
                            material_order.category = inventory_item.category
                            material_order.code = inventory_item.code
                            material_order.unit = inventory_item.unit
                        except InventoryItem.DoesNotExist:
                            material_order.name = selected_item.name
                            material_order.category = selected_item.category
                            material_order.code = selected_item.code
                            material_order.unit = selected_item.unit
                    elif selected_item:
                        material_order.name = selected_item.name
                        material_order.category = selected_item.category
                        material_order.code = selected_item.code
                        material_order.unit = selected_item.unit

                    material_order.user = request.user
                    material_order.group = request.user.groups.first() if request.user.groups.exists() else None
                    material_order.request_type = 'Receipt'
                    material_order.status = 'Draft'
                    material_order.processed_quantity = 0
                    material_order.remaining_quantity = material_order.quantity

                    # Receipt category + supplier-contract + BoQ link are now
                    # carried on MaterialOrder. ModelForm.save(commit=False)
                    # has already mapped them onto the instance — nothing
                    # extra is needed here.
                    material_order.save()

                    # If the receipt is an Overissuance Return and a BoQ line
                    # was picked, decrement that line's quantity_received so
                    # the overissuance ledger reflects the offset. The actual
                    # stock movement still happens at completion time.
                    if (
                        material_order.receipt_category == 'overissuance_return'
                        and material_order.linked_boq_item_id
                    ):
                        boq = material_order.linked_boq_item
                        # Cap so we never go negative.
                        offset = min(
                            float(material_order.quantity or 0),
                            float(boq.quantity_received or 0),
                        )
                        if offset > 0:
                            boq.quantity_received = max(
                                0.0,
                                float(boq.quantity_received or 0) - offset,
                            )
                            boq.save(update_fields=['quantity_received'])
            messages.success(request, "Material receipts submitted successfully!")
            return redirect('material_receipt')
        else:
            print("Formset errors:", formset.errors)
            messages.error(request, "There was an error with your submission.")

        # Show all inventory items to all users for transparency
        items = InventoryItem.objects.all()
        return render(request, self.template_name, {
            'formset': formset,
            'bulk_form': BulkMaterialRequestForm(),
            'items': items,
            'inventory_items': json.dumps(list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))),
            'active_tab': 'single',
            'orders': MaterialOrder.objects.filter(request_type='Receipt').order_by('-date_requested'),
        })

    def handle_bulk_receipt(self, request):
        """Handle bulk receipt uploads from Excel"""
        logger = logging.getLogger(__name__)
        
        bulk_form = BulkMaterialRequestForm(request.POST, request.FILES)
        if not bulk_form.is_valid():
            logger.error(f"Bulk form validation failed: {bulk_form.errors}")
            messages.error(request, "Bulk upload validation failed.")
            return self._render_receipt_form(request, bulk_form=bulk_form)
        
        success_count = 0
        
        try:
            df = bulk_form.cleaned_data['df']
            request_type = 'Receipt'
            priority = bulk_form.cleaned_data['priority']
            
            # Generate a base request code
            base_request_code = generate_request_code()
            df['request_code'] = [f"{base_request_code}-{i+1}" for i in range(len(df))]
            
            for idx, row in df.iterrows():
                try:
                    if not row.get('name'):
                        continue
                        
                    item_name = str(row['name']).strip()
                    item = InventoryItem.objects.filter(name__iexact=item_name).first()
                    
                    if not item:
                        logger.warning(f"Item not found: {item_name}")
                        continue
                    
                    warehouse = None
                    if 'warehouse' in row and pd.notna(row['warehouse']):
                        warehouse = Warehouse.objects.filter(name__iexact=str(row['warehouse']).strip()).first()
                    
                    with transaction.atomic():
                        order_data = {
                            'name': item.name,
                            'quantity': row['quantity'],
                            'category': item.category,
                            'code': item.code,
                            'unit': item.unit,
                            'user': request.user,
                            'group': request.user.groups.first() if request.user.groups.exists() else None,
                            'request_type': request_type,
                            'request_code': row['request_code'],
                            'warehouse': warehouse,
                            'priority': priority,
                            'status': 'Draft',
                            'processed_quantity': 0,
                            'remaining_quantity': row['quantity']
                        }
                        
                        MaterialOrder.objects.create(**order_data)
                        success_count += 1
                        
                except Exception as e:
                    logger.error(f"Error processing row {idx}: {str(e)}")
                    continue
            
            if success_count > 0:
                messages.success(request, f"Successfully created {success_count} material receipt(s)")
                return redirect('material_receipt')
                    
        except Exception as e:
            messages.error(request, f"Error processing bulk receipt: {str(e)}")
            
        return self._render_receipt_form(request, bulk_form=bulk_form)

    def _render_receipt_form(self, request, bulk_form=None):
        items = InventoryItem.objects.all()
        context = {
            'formset': MaterialReceiptFormSet(form_kwargs={'user': request.user}),
            'bulk_form': bulk_form or BulkMaterialRequestForm(),
            'items': items,
            'inventory_items': json.dumps(list(items.values('id', 'name', 'category__name', 'unit__name', 'code', 'warehouse__name'))),
            'active_tab': 'bulk' if bulk_form else 'single',
            'orders': MaterialOrder.objects.filter(request_type='Receipt').order_by('-date_requested'),
        }
        return render(request, self.template_name, context)

class MaterialReceiptListView(LoginRequiredMixin, ListView):
    template_name = 'Inventory/material_receipts.html'
    context_object_name = 'orders'

    def get_queryset(self):
        try:
            # Show all receipt orders to all users for transparency
            return MaterialOrder.objects.filter(request_type='Receipt').order_by('-date_requested')
        except Exception:
            return MaterialOrder.objects.filter(request_type='Receipt').order_by('-date_requested')


@login_required
def download_bulk_request_template(request):
    """Dynamic Excel template for the old bulk material-request flow.

    Replaces the static `bulk_request_template.xlsx` that lived under
    `static/Inventory/templates/`. Adds two batch-level columns —
    `memo_signatory_title` and `letter_signatory_title` — so the
    signatories on the generated memo and release letter can be set from
    the spreadsheet itself instead of (or in addition to) the picker on
    the bulk upload form. Only the FIRST non-empty value in each column
    is used; the rest are ignored.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("Required package openpyxl not installed.", status=500)

    import io as _io
    from django.http import HttpResponse as _HttpResponse

    columns = [
        'name', 'quantity', 'region', 'district', 'community',
        'consultant', 'contractor', 'package_number', 'warehouse',
        'memo_signatory_title', 'letter_signatory_title',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bulk requests'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    for idx, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=idx, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    width_map = {
        'name': 30, 'quantity': 12, 'region': 18, 'district': 22, 'community': 22,
        'consultant': 24, 'contractor': 24, 'package_number': 22, 'warehouse': 20,
        'memo_signatory_title': 32, 'letter_signatory_title': 32,
    }
    for idx, col in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width_map.get(col, 18)

    example = {
        'name': 'Pole - 11kV concrete', 'quantity': 24,
        'region': 'Greater Accra', 'district': 'Ga East', 'community': 'Abokobi',
        'consultant': '', 'contractor': '', 'package_number': 'SHEP-PKG-024', 'warehouse': '',
        'memo_signatory_title': 'Ag. Director, Power',
        'letter_signatory_title': 'Chief Director',
    }
    for idx, col in enumerate(columns, 1):
        ws.cell(row=2, column=idx, value=example.get(col, ''))

    ins = wb.create_sheet(title='Instructions')
    instructions = [
        "Bulk material requests (Release / Transfer)",
        "",
        "Required columns:",
        "  - name      (must match an inventory item name on record)",
        "  - quantity  (numeric, > 0)",
        "",
        "Recommended for Release:",
        "  - region, district, community, consultant, contractor, package_number, warehouse",
        "",
        "Signatory columns (batch-level — fill on the FIRST row only):",
        "  - memo_signatory_title    (Signatory title flagged for the approval memo)",
        "  - letter_signatory_title  (Signatory title flagged for the release letter)",
        "  Leave blank to use the active default set in the Signatory admin.",
        "  The picker on the upload form ALWAYS wins over the Excel value.",
        "  Unknown titles do NOT block the upload — they fall back to the default",
        "  and a warning is shown.",
        "",
        "On upload:",
        "  - Each row creates a MaterialOrder with a unique request code.",
        "  - One ReleaseLetter is created per batch (if a PDF is also attached).",
        "  - The signatory titles above populate the memo and letter signatory",
        "    fields on that ReleaseLetter.",
    ]
    for idx, line in enumerate(instructions, 1):
        cell = ins.cell(row=idx, column=1, value=line)
        if idx == 1:
            cell.font = Font(bold=True, size=14)
        elif line.endswith(':'):
            cell.font = Font(bold=True)
    ins.column_dimensions['A'].width = 90

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = _HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="bulk_request_template.xlsx"'
    return response
