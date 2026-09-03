"""Tally (bin) cards — the storekeeper-facing view of the stock ledger.

One card per InventoryItem (a material in a specific warehouse). The card shows
the running history from StockMovement, exactly like the paper bin card: date,
reference, in, out, balance. A printable PDF mirrors that paper layout.

Read-only. Movements are written by the ledger (services/stock_ledger), never
from here — this is a window onto history, not a way to edit it.
"""
import logging
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST
from django.views.generic import DetailView, ListView

from Inventory.models import InventoryItem, StockMovement, Warehouse
from Inventory.services.stock_ledger import record_movement
from Inventory.utils import can_access_stores

logger = logging.getLogger(__name__)


class StoresAccessMixin(UserPassesTestMixin):
    def test_func(self):
        return can_access_stores(self.request.user)


class StoresSupervisorMixin(UserPassesTestMixin):
    """Management / Stores Management / superuser — the integrity check is a
    supervisory tool, not a day-to-day storekeeper screen."""
    def test_func(self):
        u = self.request.user
        return u.is_superuser or u.groups.filter(
            name__in=['Management', 'Stores Management']).exists()


def _filter_movements(qs, params):
    """Apply the card's date-range and type filters from query params."""
    mtype = (params.get('type') or '').strip()
    if mtype:
        qs = qs.filter(movement_type=mtype)
    dfrom = (params.get('from') or '').strip()
    if dfrom:
        qs = qs.filter(created_at__date__gte=dfrom)
    dto = (params.get('to') or '').strip()
    if dto:
        qs = qs.filter(created_at__date__lte=dto)
    return qs


class TallyCardListView(LoginRequiredMixin, StoresAccessMixin, ListView):
    """List of stock records, each linking to its tally card."""
    model = InventoryItem
    template_name = 'Inventory/tally_card_list.html'
    context_object_name = 'items'
    paginate_by = 50

    def get_queryset(self):
        qs = (InventoryItem.objects
              .select_related('warehouse', 'unit', 'category')
              .order_by('code', 'name'))
        q = (self.request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(
                Q(code__icontains=q) | Q(name__icontains=q) |
                Q(warehouse__name__icontains=q))
        wh = (self.request.GET.get('warehouse') or '').strip()
        if wh:
            qs = qs.filter(warehouse_id=wh)
        # Low-stock only: quantity at or below a set reorder level.
        if (self.request.GET.get('low') or '') == '1':
            from django.db.models import F
            qs = qs.filter(reorder_level__gt=0, quantity__lte=F('reorder_level'))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['warehouses'] = Warehouse.objects.order_by('name')
        ctx['q'] = self.request.GET.get('q', '')
        ctx['selected_warehouse'] = self.request.GET.get('warehouse', '')
        ctx['low_only'] = (self.request.GET.get('low') or '') == '1'
        return ctx


def _card_context(item, movements=None):
    """Shared context for the on-screen card, its PDF, and Excel export."""
    if movements is None:
        movements = list(item.movements.select_related('performed_by').all())
    reorder = item.reorder_level and item.quantity <= item.reorder_level
    return {
        'item': item,
        'movements': movements,
        'reorder_flagged': bool(reorder),
        'movement_types': StockMovement.MOVEMENT_TYPES,
    }


class TallyCardDetailView(LoginRequiredMixin, StoresAccessMixin, DetailView):
    """The tally card for one stock record: header + movement history."""
    model = InventoryItem
    template_name = 'Inventory/tally_card.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        movements = list(_filter_movements(
            self.object.movements.select_related('performed_by'), self.request.GET))
        ctx.update(_card_context(self.object, movements))
        ctx['f_type'] = self.request.GET.get('type', '')
        ctx['f_from'] = self.request.GET.get('from', '')
        ctx['f_to'] = self.request.GET.get('to', '')
        return ctx


class TallyCardPDFView(LoginRequiredMixin, StoresAccessMixin, DetailView):
    """Printable bin-card PDF, styled to match the paper card. Respects the
    same date/type filters as the on-screen card."""
    model = InventoryItem

    def get(self, request, *args, **kwargs):
        item = self.get_object()
        movements = list(_filter_movements(
            item.movements.select_related('performed_by'), request.GET))
        html = render_to_string('Inventory/tally_card_print.html',
                                _card_context(item, movements))
        try:
            from weasyprint import HTML
        except Exception:  # noqa: BLE001 — renderer libs may be absent on a host
            logger.exception("WeasyPrint unavailable for tally card PDF")
            return HttpResponse(
                "PDF renderer is not available on this server. "
                "View the card on screen instead.", status=503)
        pdf = HTML(string=html).write_pdf()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'inline; filename="tally-card-{item.code or item.pk}.pdf"')
        return resp


class TallyCardExcelView(LoginRequiredMixin, StoresAccessMixin, DetailView):
    """Export a card's movements to Excel (respects the current filters)."""
    model = InventoryItem

    def get(self, request, *args, **kwargs):
        item = self.get_object()
        movements = _filter_movements(
            item.movements.select_related('performed_by'), request.GET)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Tally card'
        ws.append([f"Tally card — {item.name} ({item.code or '—'})"])
        ws.append([f"Warehouse: {item.warehouse.name if item.warehouse else '—'}   "
                   f"Unit: {item.unit.name if item.unit else '—'}   "
                   f"Balance: {item.quantity}"])
        ws.append([])
        headers = ['Date', 'Type', 'Reference', 'Received', 'Issued', 'Balance', 'By', 'Note']
        ws.append(headers)
        hfill = PatternFill('solid', fgColor='1F4E78')
        for c in range(1, len(headers) + 1):
            cell = ws.cell(4, c)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hfill
        for m in movements:
            who = ''
            if m.performed_by:
                who = m.performed_by.get_full_name() or m.performed_by.username
            ws.append([
                m.created_at.strftime('%Y-%m-%d %H:%M'),
                m.get_movement_type_display(),
                m.reference or '',
                float(m.qty_in) if m.qty_in else None,
                float(m.qty_out) if m.qty_out else None,
                float(m.balance_after),
                who or 'system',
                m.note or '',
            ])
        for i, w in enumerate([18, 20, 22, 12, 12, 14, 20, 40], 1):
            ws.column_dimensions[chr(64 + i)].width = w

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = (
            f'attachment; filename="tally-card-{item.code or item.pk}.xlsx"')
        return resp


class StockIntegrityView(LoginRequiredMixin, StoresSupervisorMixin, ListView):
    """Cards whose ledger balance has drifted from live stock — the safety net.

    A clean list here means every card's running balance equals live stock. Any
    row means something changed stock outside the ledger and needs a look.
    """
    template_name = 'Inventory/tally_card_integrity.html'
    context_object_name = 'drift_rows'

    def get_queryset(self):
        from Inventory.services.stock_ledger import find_drift
        rows = find_drift()
        return [
            {'item': it, 'live': live, 'ledger': ledger,
             'gap': (None if ledger is None else (live or 0) - ledger)}
            for it, live, ledger in rows
        ]


class TallyCardConsolidatedView(LoginRequiredMixin, StoresAccessMixin, ListView):
    """One row per material code, totalled across every warehouse it lives in."""
    template_name = 'Inventory/tally_card_consolidated.html'
    context_object_name = 'materials'
    paginate_by = 50

    def get_queryset(self):
        items = (InventoryItem.objects
                 .select_related('warehouse', 'unit')
                 .order_by('code', 'warehouse__name'))
        q = (self.request.GET.get('q') or '').strip()
        if q:
            from django.db.models import Q
            items = items.filter(Q(code__icontains=q) | Q(name__icontains=q))
        grouped = {}
        for it in items:
            key = it.code or it.name
            g = grouped.setdefault(key, {
                'code': it.code, 'name': it.name,
                'unit': it.unit.name if it.unit else '',
                'total': 0, 'locations': []})
            g['total'] += (it.quantity or 0)
            g['locations'].append({
                'pk': it.pk,
                'warehouse': it.warehouse.name if it.warehouse else '—',
                'qty': it.quantity or 0})
        return sorted(grouped.values(), key=lambda x: (x['code'] or '', x['name']))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q', '')
        return ctx


@require_POST
def tally_card_adjust(request, pk):
    """Post a physical stock count as an audited adjustment.

    The storekeeper enters the counted balance and a reason. We record the
    difference as an append-only 'adjustment' movement — history is never
    rewritten, the correction is a new line that carries who counted, when, and
    why. The item's live quantity is set to the counted figure so the ledger's
    running balance and live stock stay in lock-step.
    """
    if not can_access_stores(request.user):
        messages.error(request, "You don't have permission to adjust stock.")
        return redirect('tally_card_list')

    item = get_object_or_404(InventoryItem, pk=pk)
    reason = (request.POST.get('reason') or '').strip()
    raw = (request.POST.get('counted_quantity') or '').strip()

    if not reason:
        messages.error(request, "A reason is required for a stock adjustment.")
        return redirect('tally_card_detail', pk=pk)
    try:
        counted = Decimal(raw)
    except (InvalidOperation, ValueError):
        messages.error(request, "Enter a valid counted quantity.")
        return redirect('tally_card_detail', pk=pk)
    if counted < 0:
        messages.error(request, "Counted quantity cannot be negative.")
        return redirect('tally_card_detail', pk=pk)

    delta = counted - Decimal(item.quantity or 0)
    if delta == 0:
        messages.info(request, "Counted quantity matches the system balance — no adjustment made.")
        return redirect('tally_card_detail', pk=pk)

    item.quantity = int(counted)
    item.save(update_fields=['quantity'])
    record_movement(
        item, 'adjustment',
        qty_in=(delta if delta > 0 else 0),
        qty_out=(-delta if delta < 0 else 0),
        reference='Stock count', user=request.user,
        note=f"Adjusted to counted balance {counted}. Reason: {reason}")

    messages.success(
        request,
        f"Balance adjusted to {counted} ({'+' if delta > 0 else ''}{delta}). Logged on the card.")
    return redirect('tally_card_detail', pk=pk)
