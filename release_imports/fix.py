"""Rebuild 01 and 03 with the two faults corrected.

01: package_number is REQUIRED for SHEP and was blank -> all 43 rows rejected.
    The letter gives no package numbers, so these are DERIVED, following the
    Ministry's own format from the South Tongu letter
    (SHEP-VR-NTD-TECL-09-16 = SHEP-<region>-<district>-<contractor>-<mm>-<yy>).
    Replace with real package numbers if the Ministry has them.

03: consultant was 'Donab Engineering Limited'; prod's roster says
    'Donab Engineering'. The lookup is exact-name, so it would have bounced.
"""
from openpyxl import Workbook, load_workbook

DISTRICT_ABBR   = {'Bongo':'BON', 'Binduri':'BIN', 'Builsa North':'BUN'}
CONTRACTOR_ABBR = {
 'E. K. Asare Enterprise':'EKA',
 'Kasn Engineering and Consult Limited':'KEC',
 'Elios Engineering Limited':'EEL',
 'Jeopa Company Limited':'JCL',
 'Tirago Enterprise':'TIE',
 'Kasn And Consult Limited':'KAC',
 'Wilwak Enterprise Limited':'WEL',
 'Payaba Enterprise':'PAY',
}

def pkg(district, contractor):
    return f"SHEP-UER-{DISTRICT_ABBR[district]}-{CONTRACTOR_ABBR[contractor]}-08-26"

def sheet(path, headers, rows):
    wb = Workbook(); ws = wb.active; ws.append(headers)
    for r in rows: ws.append(r)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(14, len(h) + 4)
    wb.save(path); print(f"  {path}  ({len(rows)} rows)")

# Recover the Upper East rows from the request file (district, community, contractor).
src = load_workbook('02_requests_SHEP_upper_east_LV.xlsx').active
lv = {(r[3], r[4]): r[5] for r in src.iter_rows(min_row=2, values_only=True)}

# Full 43-row list including the HT-only communities absent from the LV file.
HT_ONLY = [('Bongo', c, 'Kasn And Consult Limited') for c in
           ['Awaa Tarongo Ext','Awale Namoo','Sikabiisi Nayari','Sikabiisi Akondone',
            'Sikabiisi Abokobisi','Lung Dangoogo','Ayourpia','Beolembisi','Beo Mooshidaboro']]
rows = [(d, c, con) for (d, c), con in lv.items()] + HT_ONLY
assert len(rows) == 43, len(rows)

sheet('01_communities_SHEP_upper_east.xlsx',
      ['region','district','community','package_number','consultant_name'],
      [['Upper East', d, c, pkg(d, con), 'Holy Engineering'] for d, c, con in rows])

# Carry the same package numbers onto the request file so the two agree.
sheet('02_requests_SHEP_upper_east_LV.xlsx',
      ['material','quantity','region','district','community','contractor','package_number','notes'],
      [[r[0], r[1], r[2], r[3], r[4], r[5], pkg(r[3], r[5]), r[6]]
       for r in src.iter_rows(min_row=2, values_only=True)])

sheet('03_communities_SHEP_south_tongu.xlsx',
      ['region','district','community','package_number','consultant_name'],
      [['Volta','South Tongu', c,'SHEP-VR-NTD-TECL-09-16','Donab Engineering']
       for c in ['Dordoekope','Adzikope','Vigbedorkope','Chiefkope/Tetsakpo']])

print('\nPackage numbers used:')
for d, c, con in sorted({(d, '', con) for d, c, con in rows}):
    print(f"   {pkg(d, con)}   {d} / {con}")
